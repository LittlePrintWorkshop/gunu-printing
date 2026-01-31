from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import json
import os
from datetime import datetime, timedelta
from functools import wraps
import jwt
import requests
import secrets
import math
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import db, User, Quote, Order, CartItem, Notice, PopupNotice, HomepageSettings, PaymentLink, CategorySettings, PaperPrice, PrintCost, PlateCost, PrintingCost, PlateCostNew, MarginSetting, BindingCost, AdditionalCost, Category, Product, ProductVariant, SellableProductOption
from PIL import Image
import tempfile
import zipfile
import base64
import subprocess

app = Flask(__name__, static_folder='.')

# DEBUG mode ON (로컬 테스트용)
app.config['DEBUG'] = True

# ========== PayApp 결제 설정 (변경 시 여기만 수정) ==========
PAYAPP_USERID = os.environ.get('PAYAPP_USERID', 'vinso112')
PAYAPP_LINKKEY = os.environ.get('PAYAPP_LINKKEY', 'RQ0pApYSGpBaGQD4VDh2ZO1DPJnCCRVaOgT+oqg6zaM=')
PAYAPP_LINKVALUE = os.environ.get('PAYAPP_LINKVALUE', 'RQ0pApYSGpBaGQD4VDh2ZKAxb4U840FF2orYsZflIx8=')
PAYAPP_CANCEL_URL = 'https://api.payapp.kr/oapi/apiLoad.html'
# ============================================================

# 데이터베이스 설정
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///printing.db')
if app.config['SQLALCHEMY_DATABASE_URI'].startswith('postgres://'):
    app.config['SQLALCHEMY_DATABASE_URI'] = app.config['SQLALCHEMY_DATABASE_URI'].replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['JWT_SECRET_KEY'] = os.environ.get('JWT_SECRET_KEY', app.config['SECRET_KEY'])
app.config['JWT_EXPIRATION_HOURS'] = 24
# 업로드 용량 제한 (20MB)
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024

# 업로드 디스크 경로 설정 (Render Persistent Disk 사용 시 /images)
UPLOAD_FOLDER = os.environ.get('UPLOAD_FOLDER', 'images')
# 절대 경로면 그대로 사용 (Render), 상대 경로면 프로젝트 폴더에 상대적으로
if not os.path.isabs(UPLOAD_FOLDER):
    UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), UPLOAD_FOLDER)

# CORS 설정 - 기본은 전체 허용(개발 편의), 배포 시 ALLOWED_ORIGINS 환경변수로 제한
allowed_origins = os.environ.get('ALLOWED_ORIGINS', '*').split(',')
CORS(app, resources={r"/api/*": {"origins": allowed_origins}})

# 캐시 방지 헤더 설정 (API 응답이 실시간으로 업데이트되도록)
@app.after_request
def set_cache_headers(response):
    # 이미지 파일은 브라우저 캐시 허용 (1개월)
    if request.path.startswith('/images/') or request.path.endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp')):
        response.headers['Cache-Control'] = 'public, max-age=2592000'  # 30일
        response.headers['Pragma'] = 'cache'
    else:
        # API 응답 및 HTML은 캐시 비활성화
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# 데이터베이스 초기화
db.init_app(app)

# ========== 계산 상수 ==========
INDIGO_CLICK = {
    'color': 200,
    'mono': 40
}
DIGITAL_CLICK = 20
OFFSET_PRICE_PER_COLOR = 8000  # 옵셋 색상당 가격 (HTML과 동기화)

# 배송비 계산용 상수
YEON_PRICE_111 = {
    "모조지": {"80": {"4x6": 73060, "kook": 50750}, "100": {"4x6": 90530, "kook": 62920}, "120": {"4x6": 108620, "kook": 75460}, "150": {"4x6": 135780, "kook": 94320}},
    "미색모조지": {"80": {"4x6": 75250, "kook": 52270}, "100": {"4x6": 93280, "kook": 64790}},
    "플러스지백색": {"80": {"4x6": 75250, "kook": 57270}, "100": {"4x6": 93240, "kook": 64790}},
    "플러스지미색": {"80": {"4x6": 77510, "kook": 53840}, "100": {"4x6": 96030, "kook": 66700}},
    "하이플러스연미": {"90": {"4x6": 88550, "kook": 61500}},
    "아트지": {"100": {"4x6": 90040, "kook": 62590}, "120": {"4x6": 108030, "kook": 75040}, "150": {"4x6": 137400, "kook": 95480}, "180": {"4x6": 164890, "kook": 114540}, "200": {"4x6": 183190, "kook": 127270}, "250": {"4x6": 228980, "kook": 159070}},
    "스노우지": {"100": {"4x6": 90040, "kook": 62590}, "120": {"4x6": 108030, "kook": 75040}, "150": {"4x6": 137400, "kook": 95480}, "180": {"4x6": 164890, "kook": 114540}, "200": {"4x6": 183190, "kook": 127270}, "250": {"4x6": 228980, "kook": 159070}}
}

# ========== 계산 함수는 이후에 정의됨 ==========

# ========== 주문 내역서 JPG 생성 함수 ==========
def create_order_invoice_html(order):
    """주문 상세 내역을 HTML로 생성"""
    try:
        # 배송정보 파싱
        delivery_info = {}
        if order.delivery_info:
            try:
                delivery_info = json.loads(order.delivery_info)
            except:
                pass
        
        # 주문상세 파싱 (금액 정보)
        order_details_list = []
        if order.order_details:
            try:
                details = json.loads(order.order_details)
                order_details_list = details if isinstance(details, list) else [details]
            except:
                pass
        
        # items 파싱 (상세 사양 정보)
        items_list = []
        if order.items:
            try:
                items = json.loads(order.items)
                items_list = items if isinstance(items, list) else [items]
            except:
                pass
        
        # 기본 정보 추출
        order_id = order.order_id or 'N/A'
        order_date = order.created_at.strftime('%Y.%m.%d %H:%M') if order.created_at else 'N/A'
        order_status = order.status or 'N/A'
        pay_type = order.pay_type or '미정'
        
        recipient = delivery_info.get('recipient', 'N/A') if delivery_info.get('recipient') else 'N/A'
        phone = delivery_info.get('phone', 'N/A') if delivery_info.get('phone') else 'N/A'
        address = delivery_info.get('address', 'N/A') if delivery_info.get('address') else 'N/A'
        
        # 금액 정보 추출
        supply_cost = 0
        vat = 0
        shipping_cost = 0
        total_price = 0
        breakdown = {}
        
        if order_details_list:
            first_item = order_details_list[0]
            supply_cost = first_item.get('supply_cost', 0)
            vat = first_item.get('vat', 0)
            shipping_cost = first_item.get('shipping', 0)
            total_price = first_item.get('total', 0)
            breakdown = first_item.get('breakdown', {})
        else:
            total_price = int(order.total_price or 0)
        
        # 상세 사양 HTML 생성
        specs_html = ""
        for item in items_list:
            item_name = item.get('name', '상품')
            item_qty = item.get('qty', 1)
            options = item.get('options', {})
            
            # 제본 방식 한글 변환
            binding_kr = {
                'staple': '중철',
                'perfect': '무선',
                'hardcover': '양장',
                'ring': '링제본'
            }
            binding_text = binding_kr.get(options.get('binding', '-'), options.get('binding', '-'))
            
            # 코팅 한글 변환
            coating_kr = {
                'none': '없음',
                'gloss': '유광',
                'matte': '무광'
            }
            coating_text = coating_kr.get(options.get('coating', '없음'), options.get('coating', '없음'))
            
            # 표지 정보 (색상 필드 제거)
            cover_html = f"""
            <div style="background:#f9f9f9; padding:12px; margin-bottom:10px; border-left:3px solid #037a3f;">
                <div style="font-weight:bold; margin-bottom:8px;">📘 표지</div>
                <div style="font-size:11px; color:#333; line-height:1.7;">
                    용지: <strong>{options.get('coverType', '-')} {options.get('coverGram', '-')}g</strong><br>
                    페이지: <strong>{options.get('coverPages', '-')}</strong><br>
                    인쇄: <strong>{options.get('coverPrint', '-')}</strong><br>
                    코팅: <strong>{coating_text}</strong>
                </div>
            </div>
            """
            
            # 내지 정보
            inner_html = f"""
            <div style="background:#f9f9f9; padding:12px; margin-bottom:10px; border-left:3px solid #0f7ba7;">
                <div style="font-weight:bold; margin-bottom:8px;">📄 내지</div>
                <div style="font-size:11px; color:#333; line-height:1.7;">
                    용지: <strong>{options.get('innerType', '-')} {options.get('innerGram', '-')}g</strong><br>
                    페이지: <strong>{options.get('innerPages', '-')}</strong><br>
                    인쇄: <strong>{options.get('innerPrint', '-')}</strong>
                </div>
            </div>
            """
            
            # 제본 정보
            binding_html = f"""
            <div style="background:#f9f9f9; padding:12px; margin-bottom:10px; border-left:3px solid #f59e0b;">
                <div style="font-weight:bold; margin-bottom:8px;">📌 제본</div>
                <div style="font-size:11px; color:#333; line-height:1.7;">
                    방식: <strong>{binding_text}</strong><br>
                    방향: <strong>{options.get('bindingDirection', '-')}</strong>
                </div>
            </div>
            """
            
            specs_html += f"""
            <div style="margin-bottom:20px; padding:15px; background:#fff; border:1px solid #ddd;">
                <div style="font-weight:bold; font-size:13px; margin-bottom:12px; color:#0f172a;">
                    {item_name}
                </div>
                <div style="margin-bottom:10px; padding:8px; background:#f8fafc;">
                    수량: <strong>{item_qty}부</strong>
                </div>
                {cover_html}
                {inner_html}
                {binding_html}
            </div>
            """
        
        # 견적 내역 HTML 생성
        breakdown_html = ""
        if breakdown:
            # 표지 비용
            cover_breakdown = breakdown.get('cover', {})
            if cover_breakdown:
                breakdown_html += f"""
                <div style="background:#fff; padding:10px; margin-bottom:10px; border:1px solid #ddd;">
                    <div style="font-weight:bold; font-size:11px; color:#037a3f; margin-bottom:6px;">표지 비용</div>
                    <div style="font-size:10px; color:#333;">
                        종이비: {cover_breakdown.get('paper', 0):,}원 | 
                        인쇄비: {cover_breakdown.get('print', 0):,}원 | 
                        판비: {cover_breakdown.get('plate', 0):,}원 | 
                        코팅비: {cover_breakdown.get('coat', 0):,}원
                    </div>
                    <div style="text-align:right; font-weight:bold; margin-top:4px;">
                        소계: {sum(cover_breakdown.values()):,}원
                    </div>
                </div>
                """
            
            # 내지 비용
            inner_breakdown = breakdown.get('inner', {})
            if inner_breakdown:
                breakdown_html += f"""
                <div style="background:#fff; padding:10px; margin-bottom:10px; border:1px solid #ddd;">
                    <div style="font-weight:bold; font-size:11px; color:#0f7ba7; margin-bottom:6px;">내지 비용</div>
                    <div style="font-size:10px; color:#333;">
                        종이비: {inner_breakdown.get('paper', 0):,}원 | 
                        인쇄비: {inner_breakdown.get('print', 0):,}원 | 
                        판비: {inner_breakdown.get('plate', 0):,}원
                    </div>
                    <div style="text-align:right; font-weight:bold; margin-top:4px;">
                        소계: {sum(inner_breakdown.values()):,}원
                    </div>
                </div>
                """
            
            # 제본 비용
            binding_cost = breakdown.get('binding', 0)
            if binding_cost:
                breakdown_html += f"""
                <div style="background:#fff; padding:10px; margin-bottom:10px; border:1px solid #ddd;">
                    <div style="font-weight:bold; font-size:11px; color:#f59e0b; margin-bottom:6px;">제본 비용</div>
                    <div style="font-size:10px; color:#333;">
                        제본비: {binding_cost:,}원
                    </div>
                </div>
                """
            
            # 배송 비용
            if shipping_cost:
                breakdown_html += f"""
                <div style="background:#fff; padding:10px; margin-bottom:10px; border:1px solid #ddd;">
                    <div style="font-weight:bold; font-size:11px; color:#8b5cf6; margin-bottom:6px;">배송 비용</div>
                    <div style="font-size:10px; color:#333;">
                        배송비: {shipping_cost:,}원
                    </div>
                </div>
                """
        
        # 간단한 상품 목록 테이블
        items_html = ""
        for idx, item in enumerate(items_list, 1):
            item_name = item.get('name', '-')
            item_qty = item.get('qty', 1)
            item_price = item.get('price', 0)
            items_html += f"""
            <tr style="border-bottom: 1px solid #ddd;">
              <td style="padding: 10px; text-align: center;">{idx}</td>
              <td style="padding: 10px; text-align: center;">{item_name}</td>
              <td style="padding: 10px; text-align: right;">{item_qty}부</td>
              <td style="padding: 10px; text-align: right;">{item_price:,}원</td>
            </tr>
            """
        
        html_content = f"""
        <!DOCTYPE html>
        <html lang="ko">
        <head>
            <meta charset="UTF-8">
            <title>주문 내역서</title>
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    color: #333;
                    margin: 0;
                    padding: 20px;
                    background: white;
                }}
                .container {{
                    max-width: 900px;
                    margin: 0 auto;
                    border: 3px solid #333;
                    padding: 40px;
                    background: #fff;
                }}
                .header {{
                    text-align: center;
                    margin-bottom: 40px;
                    border-bottom: 3px solid #333;
                    padding-bottom: 20px;
                }}
                .header h1 {{
                    margin: 0;
                    font-size: 32px;
                    font-weight: bold;
                    letter-spacing: 2px;
                }}
                .header p {{
                    margin: 5px 0 0 0;
                    color: #666;
                    font-size: 14px;
                }}
                .order-info {{
                    display: grid;
                    grid-template-columns: 1fr 1fr;
                    gap: 20px;
                    margin-bottom: 30px;
                    font-size: 13px;
                    line-height: 1.9;
                }}
                .info-box {{
                    border: 2px solid #ddd;
                    padding: 15px;
                    background: #f9f9f9;
                }}
                .info-box h3 {{
                    margin: 0 0 12px 0;
                    font-size: 14px;
                    font-weight: bold;
                    border-bottom: 2px solid #333;
                    padding-bottom: 8px;
                }}
                .info-row {{
                    display: flex;
                    justify-content: space-between;
                    margin-bottom: 6px;
                }}
                .label {{
                    font-weight: bold;
                    color: #666;
                    width: 50%;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-bottom: 25px;
                    font-size: 13px;
                }}
                th {{
                    background: #333;
                    color: white;
                    padding: 12px;
                    text-align: center;
                    font-weight: bold;
                    border: 1px solid #333;
                }}
                td {{
                    padding: 10px;
                    border: 1px solid #ddd;
                }}
                .totals {{
                    background: #f5f5f5;
                    padding: 20px;
                    border: 2px solid #ddd;
                    margin-bottom: 20px;
                }}
                .total-row {{
                    display: flex;
                    justify-content: space-between;
                    margin-bottom: 10px;
                    font-size: 14px;
                }}
                .total-row.final {{
                    font-size: 20px;
                    font-weight: bold;
                    border-top: 2px solid #333;
                    padding-top: 12px;
                    color: #d32f2f;
                }}
                .footer {{
                    text-align: center;
                    margin-top: 30px;
                    font-size: 12px;
                    color: #999;
                    border-top: 1px solid #ddd;
                    padding-top: 15px;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>주문 내역서</h1>
                    <p>Order Invoice</p>
                </div>
                
                <div class="order-info">
                    <div class="info-box">
                        <h3>주문 정보</h3>
                        <div class="info-row">
                            <span class="label">주문번호:</span>
                            <span>{order_id}</span>
                        </div>
                        <div class="info-row">
                            <span class="label">주문일:</span>
                            <span>{order_date}</span>
                        </div>
                        <div class="info-row">
                            <span class="label">상태:</span>
                            <span>{order_status}</span>
                        </div>
                        <div class="info-row">
                            <span class="label">결제:</span>
                            <span>{pay_type}</span>
                        </div>
                    </div>
                    
                    <div class="info-box">
                        <h3>배송 정보</h3>
                        <div class="info-row">
                            <span class="label">수령인:</span>
                            <span>{recipient}</span>
                        </div>
                        <div class="info-row">
                            <span class="label">연락처:</span>
                            <span>{phone}</span>
                        </div>
                        <div style="margin-top: 10px;">
                            <div class="label">배송지:</div>
                            <div style="margin-left: 0; font-size: 12px; margin-top: 4px;">
                                {address}
                            </div>
                        </div>
                    </div>
                </div>
                
                <table>
                    <thead>
                        <tr>
                            <th style="width: 10%;">번호</th>
                            <th style="width: 30%;">상품</th>
                            <th style="width: 20%; text-align: center;">수량</th>
                            <th style="width: 40%; text-align: right;">가격</th>
                        </tr>
                    </thead>
                    <tbody>
                        {items_html if items_html else '<tr><td colspan="4" style="text-align: center; padding: 20px;">상품 정보 없음</td></tr>'}
                    </tbody>
                </table>
                
                <!-- 상세 사양 -->
                {specs_html if specs_html else ''}
                
                <!-- 견적 내역 -->
                {f'''
                <div style="margin-bottom:20px; padding:15px; background:#f8fafc; border:1px solid #ddd;">
                    <div style="font-weight:bold; font-size:13px; margin-bottom:12px;">📋 견적 상세 내역</div>
                    {breakdown_html}
                </div>
                ''' if breakdown_html else ''}
                
                <div class="totals">
                    <div class="total-row">
                        <span>상품금액:</span>
                        <span>{supply_cost:,.0f}원</span>
                    </div>
                    <div class="total-row">
                        <span>배송료:</span>
                        <span>{shipping_cost:,.0f}원</span>
                    </div>
                    <div class="total-row">
                        <span>부가세 (10%):</span>
                        <span>{vat:,.0f}원</span>
                    </div>
                    <div class="total-row final">
                        <span>총 결제액:</span>
                        <span>{total_price:,.0f}원</span>
                    </div>
                </div>
                
                <div class="footer">
                    <p>발급: {datetime.now().strftime('%Y.%m.%d %H:%M:%S')}</p>
                    <p>이 문서는 전자 형태로 발급되었습니다.</p>
                </div>
            </div>
        </body>
        </html>
        """
        return html_content
    except Exception as e:
        print(f"[create_order_invoice_html] 오류: {e}")
        raise

def html_to_jpg(html_content, output_path):
    """HTML을 JPG로 변환 - 2열 레이아웃"""
    try:
        from PIL import Image, ImageDraw, ImageFont
        import re
        
        width = 1200  # 2열 레이아웃 최적화
        max_height = 3500
        img = Image.new('RGB', (width, max_height), color='white')
        draw = ImageDraw.Draw(img)
        
        # 폰트
        try:
            font_title = ImageFont.truetype("C:\\Windows\\Fonts\\malgun.ttf", 32)
            font_heading = ImageFont.truetype("C:\\Windows\\Fonts\\malgun.ttf", 15)
            font_bold = ImageFont.truetype("C:\\Windows\\Fonts\\malgun.ttf", 11)
            font_normal = ImageFont.truetype("C:\\Windows\\Fonts\\malgun.ttf", 10)
            font_small = ImageFont.truetype("C:\\Windows\\Fonts\\malgun.ttf", 9)
        except:
            font_title = font_heading = font_bold = font_normal = font_small = ImageFont.load_default()
        
        # 왼쪽 열 (50-480), 오른쪽 열 (480-1200)
        left_x = 50
        right_x = 480
        col_width = 430
        
        y_left = 40
        y_right = 40
        line_height = 24
        
        # 제목 (전체 너비)
        draw.text((50, y_left), "주문 내역서", fill='black', font=font_title)
        y_left += 60
        
        # 주문정보 추출
        order_id_match = re.search(r'주문번호:</span>\s*<span>([^<]+)</span>', html_content, re.DOTALL)
        order_id = order_id_match.group(1).strip() if order_id_match else 'N/A'
        
        created_match = re.search(r'주문일:</span>\s*<span>([^<]+)</span>', html_content, re.DOTALL)
        created_date = created_match.group(1).strip() if created_match else 'N/A'
        
        status_match = re.search(r'상태:</span>\s*<span>([^<]+)</span>', html_content, re.DOTALL)
        status = status_match.group(1).strip() if status_match else '-'
        
        pay_match = re.search(r'결제:</span>\s*<span>([^<]+)</span>', html_content, re.DOTALL)
        pay_type = pay_match.group(1).strip() if pay_match else '-'
        
        recipient_match = re.search(r'수령인:</span>\s*<span>([^<]+)</span>', html_content, re.DOTALL)
        recipient = recipient_match.group(1).strip() if recipient_match else '-'
        
        phone_match = re.search(r'연락처:</span>\s*<span>([^<]+)</span>', html_content, re.DOTALL)
        phone = phone_match.group(1).strip() if phone_match else '-'
        
        # ===== 왼쪽 열: 주문/배송/상품 정보 =====
        # 주문 정보 섹션
        draw.text((left_x, y_left), "[주문 정보]", fill='#333333', font=font_heading)
        y_left += line_height + 3
        draw.text((left_x + 15, y_left), f"주문번호: {order_id}", fill='#333333', font=font_normal)
        y_left += line_height
        draw.text((left_x + 15, y_left), f"주문일: {created_date}", fill='#333333', font=font_normal)
        y_left += line_height
        draw.text((left_x + 15, y_left), f"상태: {status} | 결제: {pay_type}", fill='#666666', font=font_small)
        y_left += int(line_height * 1.2)
        
        # 배송 정보 섹션
        draw.text((left_x, y_left), "[배송 정보]", fill='#333333', font=font_heading)
        y_left += line_height + 3
        draw.text((left_x + 15, y_left), f"수령인: {recipient}", fill='#333333', font=font_normal)
        y_left += line_height
        draw.text((left_x + 15, y_left), f"연락처: {phone}", fill='#333333', font=font_normal)
        y_left += int(line_height * 1.2)
        
        # 상품 정보 섹션
        draw.text((left_x, y_left), "[상품 정보]", fill='#333333', font=font_heading)
        y_left += line_height + 3
        
        # <table> 섹션에서 기본 상품 정보 추출
        table_match = re.search(r'<table[^>]*>(.*?)</table>', html_content, re.DOTALL)
        if table_match:
            table_html = table_match.group(1)
            rows = re.findall(r'<tr[^>]*>(.*?)</tr>', table_html, re.DOTALL)
            
            for row in rows[1:]:  # Skip header row
                cells = re.findall(r'<td[^>]*>([^<]+)</td>', row)
                if len(cells) >= 4 and cells[0].strip():
                    num, name, qty, price = cells[0].strip(), cells[1].strip(), cells[2].strip(), cells[3].strip()
                    draw.text((left_x + 15, y_left), f"{num}. {name[:40]}", fill='black', font=font_bold)
                    y_left += line_height
                    draw.text((left_x + 20, y_left), f"{qty} | {price}", fill='#666666', font=font_small)
                    y_left += int(line_height * 1.1)
        
        y_left += 10
        
        # ===== 오른쪽 열: 상세사양 정보 =====
        # 상세 사양 정보 추출
        specs_section_match = re.search(r'<!-- 상세 사양 -->.*?<!-- 견적', html_content, re.DOTALL)
        if specs_section_match:
            draw.text((right_x, y_right), "[상세 사양]", fill='#333333', font=font_heading)
            y_right += line_height + 3
            
            specs_html = specs_section_match.group(0)
            
            # 표지 정보 추출 (색상 필드 제거)
            cover_match = re.search(r'표지.*?용지:\s*<strong>([^<]+)</strong>.*?페이지:\s*<strong>([^<]+)</strong>.*?인쇄:\s*<strong>([^<]+)</strong>.*?코팅:\s*<strong>([^<]+)</strong>', specs_html, re.DOTALL)
            if cover_match:
                draw.text((right_x + 10, y_right), "표지", fill='#037a3f', font=font_bold)
                y_right += line_height
                draw.text((right_x + 20, y_right), f"용지: {cover_match.group(1).strip()}", fill='#333333', font=font_small)
                y_right += line_height - 2
                draw.text((right_x + 20, y_right), f"페이지: {cover_match.group(2).strip()}", fill='#333333', font=font_small)
                y_right += line_height - 2
                draw.text((right_x + 20, y_right), f"인쇄: {cover_match.group(3).strip()}", fill='#333333', font=font_small)
                y_right += line_height - 2
                draw.text((right_x + 20, y_right), f"코팅: {cover_match.group(4).strip()}", fill='#333333', font=font_small)
                y_right += int(line_height * 1.0)
            
            # 내지 정보 추출
            inner_match = re.search(r'내지.*?용지:\s*<strong>([^<]+)</strong>.*?페이지:\s*<strong>([^<]+)</strong>.*?인쇄:\s*<strong>([^<]+)</strong>', specs_html, re.DOTALL)
            if inner_match:
                draw.text((right_x + 10, y_right), "내지", fill='#0f7ba7', font=font_bold)
                y_right += line_height
                draw.text((right_x + 20, y_right), f"용지: {inner_match.group(1).strip()}", fill='#333333', font=font_small)
                y_right += line_height - 2
                draw.text((right_x + 20, y_right), f"페이지: {inner_match.group(2).strip()}", fill='#333333', font=font_small)
                y_right += line_height - 2
                draw.text((right_x + 20, y_right), f"인쇄: {inner_match.group(3).strip()}", fill='#333333', font=font_small)
                y_right += int(line_height * 1.0)
            
            # 제본 정보 추출
            binding_match = re.search(r'제본.*?방식:\s*<strong>([^<]+)</strong>.*?방향:\s*<strong>([^<]+)</strong>', specs_html, re.DOTALL)
            if binding_match:
                draw.text((right_x + 10, y_right), "제본", fill='#f59e0b', font=font_bold)
                y_right += line_height
                draw.text((right_x + 20, y_right), f"방식: {binding_match.group(1).strip()}", fill='#333333', font=font_small)
                y_right += line_height - 2
                draw.text((right_x + 20, y_right), f"방향: {binding_match.group(2).strip()}", fill='#333333', font=font_small)
                y_right += int(line_height * 1.2)
        
        # 최대 높이 계산
        max_y = max(y_left, y_right) + 50
        
        # 구분선 (전체 너비)
        draw.line([(30, max_y), (1180, max_y)], fill='#999999', width=1)
        max_y += 15
        
        # 결제 정보 (하단에 전체 너비로)
        draw.text((left_x, max_y), "[결제 정보]", fill='#333333', font=font_heading)
        max_y += line_height + 3
        
        supply_match = re.search(r'상품금액:</span>\s*<span>([^<]+)</span>', html_content, re.DOTALL)
        if supply_match:
            draw.text((left_x + 15, max_y), f"상품금액: {supply_match.group(1).strip()}", fill='#333333', font=font_normal)
            max_y += line_height
        
        shipping_match = re.search(r'배송료:</span>\s*<span>([^<]+)</span>', html_content, re.DOTALL)
        if shipping_match:
            draw.text((left_x + 15, max_y), f"배송료: {shipping_match.group(1).strip()}", fill='#333333', font=font_normal)
            max_y += line_height
        
        vat_match = re.search(r'부가세.*?<span>([^<]+)</span>', html_content, re.DOTALL)
        if vat_match:
            draw.text((left_x + 15, max_y), f"부가세: {vat_match.group(1).strip()}", fill='#333333', font=font_normal)
            max_y += line_height
        
        total_match = re.search(r'총 결제액:</span>\s*<span>([^<]+)</span>', html_content, re.DOTALL)
        if total_match:
            max_y += 5
            draw.line([(30, max_y), (1180, max_y)], fill='#333333', width=2)
            max_y += 10
            draw.text((left_x + 15, max_y), f"총 결제액: {total_match.group(1).strip()}", fill='red', font=font_heading)
            max_y += line_height + 20
        
        # 실제 필요한 높이와 너비로 자동 조절
        # 오른쪽 최대값: right_x + 350 정도
        final_height = max_y
        final_width = 900  # 오른쪽 여백 제거
        
        # 항상 crop 수행
        img = img.crop((0, 0, final_width, final_height))
        
        img.save(output_path, 'JPEG', quality=90)
        print(f"JPG created: {output_path}")
        
    except Exception as e:
        print(f"[html_to_jpg] Error: {e}")
        import traceback
        traceback.print_exc()
        raise


# ========== 이미지 최적화 함수 ==========
def optimize_image(image_path, output_format='webp'):
    """
    이미지 압축 및 WebP 변환
    - 이미지 크기: 최대 1920x1440 (가로x세로)
    - 퀄리티: 85 (손실 압축)
    - 포맷: WebP (더 작은 용량)
    """
    try:
        # 원본 이미지 열기
        img = Image.open(image_path)
        
        # RGB 모드로 변환 (RGBA, CMYK 등 호환성)
        if img.mode in ('RGBA', 'LA', 'P'):
            # 투명도가 있는 경우 배경색(흰색) 추가
            background = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
            img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        
        # 이미지 크기 조정 (최대 1920x1440)
        max_width, max_height = 1920, 1440
        img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
        
        # WebP로 저장 (85% 품질)
        output_path = os.path.splitext(image_path)[0] + '.webp'
        img.save(output_path, 'WEBP', quality=85, method=6)
        
        # 원본 파일 삭제
        if os.path.exists(image_path) and image_path != output_path:
            os.remove(image_path)
        
        return output_path, 'webp'
    except Exception as e:
        print(f"이미지 최적화 오류: {e}")
        return image_path, os.path.splitext(image_path)[1][1:].lower()

# ========== 계산 함수는 이후에 정의됨 ==========

# 테이블 생성
with app.app_context():
    db.create_all()
    
    # 기존 DB 마이그레이션: favicon 컬럼 추가 (없을 경우)
    try:
        from sqlalchemy import inspect, text
        inspector = inspect(db.engine)
        
        # homepage_settings 테이블 확인
        if 'homepage_settings' in inspector.get_table_names():
            columns = [c['name'] for c in inspector.get_columns('homepage_settings')]
            if 'favicon' not in columns:
                # PostgreSQL 또는 SQLite 맞게 처리
                db.session.execute(text('ALTER TABLE homepage_settings ADD COLUMN favicon VARCHAR(500)'))
                db.session.commit()
                print("✅ favicon 컬럼 추가됨")

        # orders 테이블에 order_details 컬럼 없으면 추가
        if 'orders' in inspector.get_table_names():
            order_cols = [c['name'] for c in inspector.get_columns('orders')]
            if 'order_details' not in order_cols:
                db.session.execute(text('ALTER TABLE orders ADD COLUMN order_details TEXT'))
                db.session.commit()
                print("✅ orders.order_details 컬럼 추가됨")
            
            # PayApp 거래번호 필드 추가
            if 'mul_no' not in order_cols:
                db.session.execute(text('ALTER TABLE orders ADD COLUMN mul_no VARCHAR(100)'))
                db.session.commit()
                print("✅ orders.mul_no 컬럼 추가됨")
            
            if 'pay_type' not in order_cols:
                db.session.execute(text('ALTER TABLE orders ADD COLUMN pay_type VARCHAR(50)'))
                db.session.commit()
                print("✅ orders.pay_type 컬럼 추가됨")
            
            # 택배 송장 번호 필드 추가
            if 'tracking_number' not in order_cols:
                db.session.execute(text('ALTER TABLE orders ADD COLUMN tracking_number VARCHAR(100)'))
                db.session.commit()
                print("✅ orders.tracking_number 컬럼 추가됨")
        
        # category_settings 테이블 확인 및 생성
        if 'category_settings' not in inspector.get_table_names():
            db.session.execute(text('''
                CREATE TABLE category_settings (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    category VARCHAR(50) UNIQUE NOT NULL,
                    settings_data TEXT NOT NULL,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            '''))
            db.session.commit()
            print("✅ category_settings 테이블 생성됨")
        
        # 신규 테이블: categories (견적형/판매형 카테고리)
        if 'categories' not in inspector.get_table_names():
            db.session.execute(text('''
                CREATE TABLE categories (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name VARCHAR(100) NOT NULL,
                    description TEXT,
                    category_type VARCHAR(20) NOT NULL,
                    icon VARCHAR(100),
                    display_order INTEGER DEFAULT 0,
                    is_active BOOLEAN DEFAULT 1,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            '''))
            db.session.commit()
            print("✅ categories 테이블 생성됨")
        
        # 신규 테이블: products (견적형/판매형 상품)
        if 'products' not in inspector.get_table_names():
            db.session.execute(text('''
                CREATE TABLE products (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    category_id INTEGER NOT NULL,
                    name VARCHAR(200) NOT NULL,
                    description TEXT,
                    product_type VARCHAR(20) NOT NULL,
                    margin INTEGER DEFAULT 0,
                    image_url VARCHAR(500),
                    is_active BOOLEAN DEFAULT 1,
                    display_order INTEGER DEFAULT 0,
                    quote_settings TEXT,
                    fixed_price REAL,
                    quantity INTEGER DEFAULT 0,
                    stock_alert INTEGER DEFAULT 10,
                    cost_price REAL,
                    sellable_specs TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(category_id) REFERENCES categories(id)
                )
            '''))
            db.session.commit()
            print("✅ products 테이블 생성됨")
        
        # 신규 테이블: product_variants (제본/옵션)
        if 'product_variants' not in inspector.get_table_names():
            db.session.execute(text('''
                CREATE TABLE product_variants (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER NOT NULL,
                    binding_type VARCHAR(50),
                    guide_text TEXT,
                    ship_info TEXT,
                    info_html TEXT,
                    variant_price REAL,
                    variant_specs TEXT,
                    display_order INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(product_id) REFERENCES products(id)
                )
            '''))
            db.session.commit()
            print("✅ product_variants 테이블 생성됨")
        
        # 신규 테이블: sellable_product_options (옵션)
        if 'sellable_product_options' not in inspector.get_table_names():
            db.session.execute(text('''
                CREATE TABLE sellable_product_options (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER NOT NULL,
                    option_name VARCHAR(100) NOT NULL,
                    option_values TEXT NOT NULL,
                    is_required BOOLEAN DEFAULT 0,
                    display_order INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(product_id) REFERENCES products(id)
                )
            '''))
            db.session.commit()
            print("✅ sellable_product_options 테이블 생성됨")
    except Exception as e:
        print(f"⚠️ 마이그레이션 처리: {e}")
        db.session.rollback()
    
    # 기본 관리자 계정 생성 (없을 경우)
    admin_user = User.query.filter_by(user_id='admin').first()
    if not admin_user:
        admin_user = User(user_id='admin', name='관리자', role='admin', email='')
        admin_user.set_password(os.environ.get('ADMIN_DEFAULT_PASSWORD', 'admin1234'))
        db.session.add(admin_user)
        db.session.commit()

# ========== JWT 유틸리티 함수 ==========
def create_token(user_id):
    """JWT 토큰 생성"""
    payload = {
        'user_id': user_id,
        'exp': datetime.utcnow() + timedelta(hours=app.config['JWT_EXPIRATION_HOURS']),
        'iat': datetime.utcnow()
    }
    return jwt.encode(payload, app.config['JWT_SECRET_KEY'], algorithm='HS256')

def verify_token(token):
    """JWT 토큰 검증"""
    try:
        payload = jwt.decode(token, app.config['JWT_SECRET_KEY'], algorithms=['HS256'])
        return payload['user_id']
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None

def token_required(f):
    """JWT 인증 데코레이터 (로컬 환경에서는 선택적)"""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        is_local = request.host.startswith('127.0.0.1') or request.host.startswith('localhost')
        
        # Authorization 헤더에서 토큰 추출
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                token = auth_header.split(' ')[1]  # "Bearer <token>"
            except IndexError:
                return jsonify({'success': False, 'message': '잘못된 토큰 형식입니다.'}), 401
        
        if not token:
            # 로컬 환경이면 테스트 사용자로 진행
            if is_local:
                print(f"[token_required] 🔧 로컬 환경에서 토큰 없음 - 테스트 모드로 진행")
                test_user = User.query.filter_by(user_id='test_user').first()
                if not test_user:
                    # 테스트 사용자 자동 생성
                    test_user = User(user_id='test_user', name='Test User', role='user')
                    test_user.set_password('test')
                    db.session.add(test_user)
                    db.session.commit()
                    print(f"[token_required] ✅ 테스트 사용자 생성됨")
                return f(test_user, *args, **kwargs)
            else:
                return jsonify({'success': False, 'message': '토큰이 필요합니다.'}), 401
        
        user_id = verify_token(token)
        if not user_id:
            return jsonify({'success': False, 'message': '유효하지 않거나 만료된 토큰입니다.'}), 401
        
        # 사용자 정보 조회
        current_user = User.query.filter_by(user_id=user_id).first()
        if not current_user:
            return jsonify({'success': False, 'message': '사용자를 찾을 수 없습니다.'}), 401
        
        return f(current_user, *args, **kwargs)
    
    return decorated

def admin_required(f):
    """관리자 권한 데코레이터"""
    @wraps(f)
    def decorated(current_user, *args, **kwargs):
        if current_user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        return f(current_user, *args, **kwargs)
    
    return decorated

def _cancel_payapp_payment(mul_no, total_price, cancel_memo):
    """PayApp 결제 취소 함수"""
    try:
        # 취소 요청 파라미터
        cancel_params = {
            'cmd': 'paycancel',
            'userid': PAYAPP_USERID,
            'linkkey': PAYAPP_LINKKEY,
            'mul_no': mul_no,
            'cancelmemo': cancel_memo,
            'partcancel': '0'  # 0 = 전액취소
        }
        
        # POST 요청
        response = requests.post(PAYAPP_CANCEL_URL, data=cancel_params, timeout=10)
        response_text = response.text
        
        print(f"📊 PayApp 취소 응답: {response_text}")
        
        # 응답 파싱 (PayApp은 텍스트 또는 JSON으로 응답)
        try:
            # JSON 형식으로 시도
            response_data = response.json() if response.headers.get('content-type', '').count('application/json') > 0 else {}
        except:
            # 텍스트 응답 파싱
            response_data = {}
            if 'state=1' in response_text:
                response_data['state'] = '1'
            elif 'errorMessage' in response_text:
                parts = response_text.split('errorMessage=')
                if len(parts) > 1:
                    response_data['errorMessage'] = parts[1].split('&')[0]
        
        # 취소 성공 여부 확인
        success = response_data.get('state') == '1' or 'state=1' in response_text
        
        if success:
            return {
                'success': True,
                'message': '결제가 취소되었습니다.',
                'mul_no': mul_no
            }
        else:
            error_msg = response_data.get('errorMessage', '알 수 없는 오류')
            return {
                'success': False,
                'error': error_msg
            }
    
    except requests.exceptions.Timeout:
        return {
            'success': False,
            'error': 'PayApp 서버 응답 시간 초과'
        }
    except requests.exceptions.RequestException as e:
        return {
            'success': False,
            'error': f'네트워크 오류: {str(e)}'
        }
    except Exception as e:
        print(f"❌ PayApp 취소 중 오류: {e}")
        return {
            'success': False,
            'error': str(e)
        }

# 정적 파일 서빙
@app.route('/', methods=['GET', 'POST'])
def index():
    return send_from_directory('.', 'index.html')

@app.route('/payment-complete.html', methods=['GET', 'POST', 'HEAD'])
def payment_complete():
    """결제 완료 페이지"""
    return send_from_directory('.', 'payment-complete.html', mimetype='text/html')

@app.route('/api/payment-callback', methods=['POST'])
def payment_callback():
    """PayApp 결제 완료 콜백 - feedbackurl"""
    try:
        # PayApp에서 POST로 전송되는 결제 정보 받기 (form 데이터)
        data = request.form.to_dict()
        print(f"📡 PayApp 콜백 수신 (전체): {data}")
        
        # 필수 정보 추출
        state = data.get('state')  # 1이면 성공
        mul_no = data.get('mul_no')  # 결제요청번호
        pay_type = data.get('pay_type')  # 결제 타입
        order_id = data.get('var1')  # 주문번호 (PayApp 요청 시 var1에 넣음)
        
        print(f"🔍 결제 상태: state={state}, mul_no={mul_no}, order_id={order_id}, pay_type={pay_type}")
        
        # mul_no와 order_id가 있으면 저장 시도
        if mul_no and order_id:
            try:
                # 주문 찾기
                order = Order.query.filter_by(order_id=order_id).first()
                if order:
                    # [Fix] mul_no, pay_type 저장
                    order.mul_no = mul_no
                    order.pay_type = pay_type
                    
                    # [Fix] pay_type이 있을 때만 상태를 '주문접수'로 변경
                    # pay_type이 없으면 = 아직 미결제 상태 유지 (프론트에서 결제가 완료되지 않았다는 뜻)
                    if pay_type:
                        order.status = '주문접수'
                        print(f"✅ 주문 {order_id}에 mul_no={mul_no}, pay_type={pay_type}, status=주문접수 저장 완료")
                    else:
                        print(f"⚠️ 주문 {order_id}에 mul_no={mul_no} 저장만 (pay_type 미수신 - 상태 유지)")
                    
                    db.session.commit()
                else:
                    print(f"⚠️ 주문을 찾을 수 없음: {order_id} (나중에 업데이트될 수도 있음)")
            except Exception as e:
                print(f"⚠️ mul_no 저장 중 에러 (무시하고 계속): {e}")
        else:
            print(f"⚠️ mul_no 또는 order_id 없음 (무시)")
        
        # PayApp에 무조건 성공 응답 반환 (200 OK)
        return 'OK', 200
            
    except Exception as e:
        print(f"❌ PayApp 콜백 처리 에러: {e}")
        import traceback
        traceback.print_exc()
        # 에러가 발생해도 OK 반환 (PayApp 재시도 방지)
        return 'OK', 200

# [Debug] 로컬 테스트용 - 결제대기 주문을 주문접수로 변경
@app.route('/payment-complete-close', methods=['GET', 'POST'])
def payment_complete_close():
    """PayApp returnurl - 팝업 닫기 신호 전송"""
    print("[payment_complete_close] PayApp에서 리다이렉트됨 - 팝업 닫기")
    
    # PayApp 팝업 내에서 실행되는 페이지
    # opener(부모 창)에 메시지를 보내서 팝업을 닫도록 함
    html = '''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>결제 완료</title>
    <link rel="icon" href="data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHZpZXdCb3g9IjAgMCA1MTIgNTEyIj48cmVjdCB3aWR0aD0iNTEyIiBoZWlnaHQ9IjUxMiIgZmlsbD0iIzI1NjNlYiIvPjx0ZXh0IHg9IjI1NiIgeT0iMzgwIiBmb250LXNpemU9IjI4MCIgZm9udC13ZWlnaHQ9ImJvbGQiIGZpbGw9IndoaXRlIiB0ZXh0LWFuY2hvcj0ibWlkZGxlIiBmb250LWZhbWlseT0iQXJpYWwsIHNhbnMtc2VyaWYiPkc8L3RleHQ+PC9zdmc+" type="image/svg+xml">
    <style>
        body { font-family: Arial, sans-serif; margin: 0; padding: 20px; text-align: center; background: #f0f0f0; }
        p { color: #333; font-size: 16px; }
    </style>
</head>
<body>
    <p>✅ 결제가 완료되었습니다.</p>
    <p style="font-size: 12px; color: #666;">창을 닫는 중입니다...</p>
    <script>
        console.log('[payment_complete_close] 페이지 로드됨');
        
        // 부모 창에 "팝업 닫아달라" 신호 전송
        if (window.opener && !window.opener.closed) {
            console.log('[payment_complete_close] opener 감지됨 - 팝업 종료 신호 전송');
            try {
                // 1. 부모 창에 payappWindow 변수 초기화 신호 (monitorPaymentWindow 중지)
                window.opener.postMessage({
                    type: 'payment_completed_from_payapp',
                    message: 'PayApp에서 결제 완료됨',
                    closePopup: true
                }, '*');
                console.log('[payment_complete_close] 신호 전송 완료');
            } catch (e) {
                console.error('[payment_complete_close] 신호 전송 실패:', e);
            }
        }
        
        // 2. 팝업 자신을 종료하려고 시도 (작동 보장 불가)
        function closeWindow() {
            try {
                window.close();
            } catch (e) {
                console.log('[payment_complete_close] window.close() 실패');
            }
        }
        
        closeWindow();
        setTimeout(closeWindow, 100);
        setTimeout(closeWindow, 300);
        setTimeout(closeWindow, 700);
        
        // 3. 1초 후에도 안 닫혔으면 홈페이지로 강제 이동
        setTimeout(() => {
            if (!window.closed) {
                console.log('[payment_complete_close] 팝업 닫기 실패 - 홈페이지로 이동');
                try {
                    window.location.href = '/';
                } catch (e) {}
            }
        }, 1500);
    </script>
</body>
</html>'''
    
    return html

@app.route('/<path:path>', methods=['GET', 'POST', 'HEAD', 'OPTIONS'])
def static_files(path):
    # [Fix] /api/ 로 시작하면 처리하지 않음 (API 라우트에 맡김)
    if path.startswith('api/'):
        return '', 404
    
    # [Fix] 특정 라우트는 처리하지 않음 (Flask가 처리하도록)
    if path in ['login', 'signup']:
        return '', 404
    
    # JavaScript 파일의 MIME type을 명시적으로 설정
    if path.endswith('.js'):
        return send_from_directory('.', path, mimetype='application/javascript')
    elif path.endswith('.css'):
        return send_from_directory('.', path, mimetype='text/css')
    elif path.endswith('.html'):
        return send_from_directory('.', path, mimetype='text/html')
    return send_from_directory('.', path)

@app.route('/images/<path:filename>')
def serve_images(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

# ========== 로컬 테스트 헬퍼 API ==========
@app.route('/api/test/auto-update-order-status', methods=['POST'])
def test_auto_update_order_status():
    """[로컬 테스트용] 주문 상태를 자동으로 다음 단계로 업데이트"""
    # 로컬 환경에서만 작동
    is_local = request.host.startswith('127.0.0.1') or request.host.startswith('localhost')
    if not is_local:
        return jsonify({'success': False, 'message': '로컬 환경에서만 사용 가능합니다.'}), 403
    
    try:
        data = request.json
        order_id = data.get('order_id')
        
        order = Order.query.filter_by(order_id=order_id).first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
        
        # 상태 전환 로직
        status_flow = {
            '결제대기': '주문접수',
            '주문접수': '제작중',
            '제작중': '배송중',
            '배송중': '배송완료'
        }
        
        old_status = order.status
        new_status = status_flow.get(old_status)
        
        if not new_status:
            return jsonify({
                'success': False, 
                'message': f'상태 \'{old_status}\'에서 업데이트할 수 없습니다. (최종 상태 또는 잘못된 상태)'
            }), 400
        
        # 상태 업데이트
        order.status = new_status
        db.session.commit()
        
        print(f"[테스트] 주문 상태 자동 업데이트: {order_id} ({old_status} → {new_status})")
        
        return jsonify({
            'success': True,
            'message': f'상태 업데이트 완료',
            'order_id': order_id,
            'old_status': old_status,
            'new_status': new_status
        })
    
    except Exception as e:
        print(f"[ERROR] 테스트 상태 업데이트 실패: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/test/auto-accept-order', methods=['POST'])
def test_auto_accept_order():
    """[로컬 테스트용] 결제대기 주문을 주문접수로 자동 변경"""
    # 로컬 환경에서만 작동
    is_local = request.host.startswith('127.0.0.1') or request.host.startswith('localhost')
    if not is_local:
        return jsonify({'success': False, 'message': '로컬 환경에서만 사용 가능합니다.'}), 403
    
    try:
        data = request.json
        order_id = data.get('order_id')
        
        order = Order.query.filter_by(order_id=order_id).first()
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
        
        if order.status != '결제대기':
            return jsonify({
                'success': False,
                'message': f'이미 접수된 주문입니다. (현재 상태: {order.status})'
            }), 400
        
        # 결제대기 → 주문접수
        order.status = '주문접수'
        db.session.commit()
        
        print(f"[테스트] 주문 자동 접수: {order_id} (결제대기 → 주문접수)")
        
        return jsonify({
            'success': True,
            'message': '주문이 자동으로 접수되었습니다.',
            'order_id': order_id,
            'status': order.status
        })
    
    except Exception as e:
        print(f"[ERROR] 주문 자동 접수 실패: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ========== 사용자 관련 API ==========
@app.route('/api/users/login', methods=['POST'])
def login():
    data = request.json
    user = User.query.filter_by(user_id=data.get('id')).first()
    
    if user and user.check_password(data.get('pw')):
        token = create_token(user.user_id)
        return jsonify({
            'success': True,
            'token': token,
            'user': user.to_dict()
        })
    else:
        return jsonify({'success': False, 'message': '아이디 또는 비밀번호가 일치하지 않습니다.'})

@app.route('/api/users/signup', methods=['POST'])
def signup():
    data = request.json
    
    # 중복 확인
    if User.query.filter_by(user_id=data.get('id')).first():
        return jsonify({'success': False, 'message': '이미 사용 중인 아이디입니다.'})
    
    new_user = User(
        user_id=data.get('id'),
        name=data.get('name'),
        email=data.get('email', ''),
        phone=data.get('phone', ''),
        company=data.get('company', ''),
        address=data.get('address', ''),
        role='user'
    )
    new_user.set_password(data.get('pw'))
    
    db.session.add(new_user)
    db.session.commit()
    
    return jsonify({'success': True, 'message': '회원가입이 완료되었습니다.'})

@app.route('/api/users/check-id', methods=['POST'])
def check_id():
    data = request.json
    exists = User.query.filter_by(user_id=data.get('id')).first() is not None
    return jsonify({'available': not exists})

# 현재 사용자 정보 조회 (개인결제 링크용)
@app.route('/api/user/profile', methods=['GET'])
@token_required
def get_user_profile(current_user):
    """현재 로그인한 사용자의 프로필 정보 조회"""
    return jsonify({
        'success': True,
        'user': current_user.to_dict()
    })

# 회원 목록 조회 (관리자 전용)
@app.route('/api/users', methods=['GET'])
def get_users():
    """사용자 목록 조회 (관리자용, 로컬 환경에서는 토큰 없이도 가능)"""
    # 토큰 확인 및 사용자 결정
    token = None
    current_user = None
    is_local = request.host.split(':')[0] in ['localhost', '127.0.0.1'] or 'localhost' in request.host
    
    if 'Authorization' in request.headers:
        try:
            auth_header = request.headers['Authorization']
            token = auth_header.split(' ')[1]
            user_id = verify_token(token)
            if user_id:
                current_user = User.query.filter_by(user_id=user_id).first()
        except:
            pass
    
    # 토큰이 없으면 로컬 환경에서는 테스트 관리자 사용
    if not current_user:
        if is_local:
            admin_user = User.query.filter_by(user_id='admin', role='admin').first()
            if not admin_user:
                admin_user = User(user_id='admin', name='Admin', role='admin')
                admin_user.set_password('admin')
                db.session.add(admin_user)
                db.session.commit()
            current_user = admin_user
        else:
            return jsonify({'success': False, 'message': '토큰이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': '권한이 없습니다.'}), 403
    
    users = User.query.all()
    users_list = [{
        'db_id': u.id,
        'user_id': u.user_id,
        'name': u.name,
        'email': u.email,
        'phone': u.phone,
        'addr': u.address,
        'addr_detail': '',
        'role': u.role,
        'biz_name': u.company if u.role == 'business' else '',
        'biz_num': '',
        'status': 'active',
        'created_at': u.created_at.isoformat() if u.created_at else None
    } for u in users]
    
    return jsonify({'success': True, 'users': users_list})

# 회원 정보 수정
@app.route('/api/users/<user_id>', methods=['PUT'])
@token_required
def update_user(current_user, user_id):
    # 본인 또는 관리자만 수정 가능
    if current_user.user_id != user_id and current_user.role != 'admin':
        return jsonify({'success': False, 'message': '권한이 없습니다.'}), 403
    
    user = User.query.filter_by(user_id=user_id).first()
    if not user:
        return jsonify({'success': False, 'message': '사용자를 찾을 수 없습니다.'}), 404
    
    data = request.json
    
    # 비밀번호 변경 요청이 있는 경우
    if 'current_pw' in data and 'new_pw' in data:
        if not user.check_password(data['current_pw']):
            return jsonify({'success': False, 'message': '현재 비밀번호가 일치하지 않습니다.'}), 400
        user.set_password(data['new_pw'])
    
    # 기타 정보 업데이트
    if 'name' in data:
        user.name = data['name']
    if 'phone' in data:
        user.phone = data['phone']
    if 'addr' in data:
        user.address = data['addr']
    if 'email' in data:
        user.email = data['email']
    if 'biz_name' in data:
        user.company = data['biz_name']
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': '회원정보가 수정되었습니다.'})

# 회원 삭제 (관리자 전용)
@app.route('/api/users/<user_id>', methods=['DELETE'])
@token_required
def delete_user(current_user, user_id):
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': '권한이 없습니다.'}), 403
    
    if user_id == 'admin':
        return jsonify({'success': False, 'message': '관리자 계정은 삭제할 수 없습니다.'}), 400
    
    user = User.query.filter_by(user_id=user_id).first()
    if not user:
        return jsonify({'success': False, 'message': '사용자를 찾을 수 없습니다.'}), 404
    
    db.session.delete(user)
    db.session.commit()
    
    return jsonify({'success': True, 'message': '회원이 삭제되었습니다.'})

# ========== 견적 관련 API ==========
@app.route('/api/quotes', methods=['GET'])
@token_required
def get_quotes(current_user):
    quotes = Quote.query.filter_by(user_db_id=current_user.id).all()
    
    return jsonify({'success': True, 'quotes': [q.to_dict() for q in quotes]})

@app.route('/api/quotes', methods=['POST'])
@token_required
def create_quote(current_user):
    data = request.json
    
    quote_id = f"Q{datetime.now().strftime('%Y%m%d%H%M%S')}"
    new_quote = Quote(
        quote_id=quote_id,
        user_db_id=current_user.id,
        category=data.get('category'),
        binding=data.get('binding'),
        specs=json.dumps(data.get('specs', {}), ensure_ascii=False),
        price=data.get('price'),
        quote_details=json.dumps(data.get('quote_details', {}), ensure_ascii=False)
    )
    
    db.session.add(new_quote)
    db.session.commit()
    
    return jsonify({'success': True, 'quote_id': quote_id})

@app.route('/api/quotes/<quote_id>', methods=['GET'])
@token_required
def get_quote(current_user, quote_id):
    quote = Quote.query.filter_by(quote_id=quote_id).first()
    
    if quote and quote.user_db_id == current_user.id:
        return jsonify({'success': True, 'quote': quote.to_dict()})
    else:
        return jsonify({'success': False, 'message': '견적을 찾을 수 없거나 권한이 없습니다.'})

# ========== 주문 관련 API ==========
@app.route('/api/orders', methods=['GET'])
def get_orders():
    """사용자의 주문 목록 조회 (로컬 환경에서는 토큰 없이도 가능)"""
    token = None
    current_user = None
    is_local = request.host.split(':')[0] in ['localhost', '127.0.0.1'] or 'localhost' in request.host
    
    if 'Authorization' in request.headers:
        try:
            auth_header = request.headers['Authorization']
            token = auth_header.split(' ')[1]
            user_id = verify_token(token)
            if user_id:
                current_user = User.query.filter_by(user_id=user_id).first()
        except:
            pass
    
    if not current_user:
        if is_local:
            test_user = User.query.filter_by(user_id='test_user').first()
            if not test_user:
                test_user = User(user_id='test_user', name='Test User', role='user')
                test_user.set_password('test')
                db.session.add(test_user)
                db.session.commit()
            current_user = test_user
        else:
            return jsonify({'success': False, 'message': '토큰이 필요합니다.'}), 401
    
    # 모든 주문 조회 (결제대기 상태 제외, 주문접수 이상만)
    orders = Order.query.filter_by(user_db_id=current_user.id).filter(
        Order.status != '결제대기'
    ).order_by(Order.created_at.desc()).all()
    
    return jsonify({'success': True, 'orders': [o.to_dict() for o in orders]})

@app.route('/api/orders', methods=['POST'])
def create_order():
    try:
        # 토큰 확인 및 사용자 결정
        token = None
        current_user = None
        is_local = request.host.split(':')[0] in ['localhost', '127.0.0.1'] or 'localhost' in request.host
        
        if 'Authorization' in request.headers:
            try:
                auth_header = request.headers['Authorization']
                token = auth_header.split(' ')[1]
                user_id = verify_token(token)
                if user_id:
                    current_user = User.query.filter_by(user_id=user_id).first()
            except:
                pass
        
        # 토큰이 없으면 로컬 환경에서는 테스트 사용자 사용
        if not current_user:
            if is_local:
                print(f"[POST /api/orders] 🔧 로컬 환경 - 테스트 사용자 사용")
                test_user = User.query.filter_by(user_id='test_user').first()
                if not test_user:
                    test_user = User(user_id='test_user', name='Test User', role='user')
                    test_user.set_password('test')
                    db.session.add(test_user)
                    db.session.commit()
                current_user = test_user
            else:
                return jsonify({'success': False, 'message': '토큰이 필요합니다.'}), 401
        
        data = request.json
        print(f"[POST /api/orders] 🔧 주문 생성 시작 - user_id={current_user.id}")
        print(f"[POST /api/orders] 📦 받은 데이터: {data}")
        
        # 주문번호: O고객번호-날짜시간 (예: O00001-20260121123456)
        customer_num = f"{current_user.id:05d}"
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        order_id = f"O{customer_num}-{timestamp}"
        customer_code = f"C{current_user.id:05d}"
        
        payment_info = data.get('payment_info', {})
        items = data.get('items', [])
        
        # 각 아이템별로 백엔드 계산 수행
        order_details_list = []
        total_price = 0
        
        for item in items:
            category = item.get('category')
            qty = item.get('qty', 0)
            specs = item.get('specs', {})
            member_type = data.get('member_type', 'general')
            
            # qty를 정수로 변환 (문자열이면 숫자만 추출)
            if isinstance(qty, str):
                import re
                qty_match = re.search(r'\d+', qty)
                qty = int(qty_match.group()) if qty_match else 0
            else:
                qty = int(qty) if qty else 0
            
            # category 유효성 검사
            if not category:
                print(f"[POST /api/orders] ⚠️ category가 없음: {item}")
                raise ValueError(f"카테고리가 없습니다: {item}")
            
            try:
                # 백엔드에서 계산 수행
                if category.startswith('flyer'):
                    calc_result = calculate_flyer_backend(category, qty, 0, {}, specs)
                else:  # indigo, digital, offset
                    calc_result = calculate_book_backend(category, qty, 0, {}, specs)
                
                # 마진 적용 (DB에서 조회)
                margin = get_margin_from_db(category, member_type)
                calc_result_with_margin = {
                    **calc_result,
                    'margin_rate': margin
                }
                
                # 마진을 다시 적용해서 최종 가격 계산
                if margin > 0:
                    supply_cost = int(calc_result['supply_cost'] * (1 + margin / 100))
                    vat = int(supply_cost * 0.1)
                    final_price = int((supply_cost + vat) / 10) * 10
                else:
                    supply_cost = calc_result['supply_cost']
                    vat = calc_result['vat']
                    final_price = calc_result['total']
                
                calc_result_with_margin['supply_cost'] = supply_cost
                calc_result_with_margin['vat'] = vat
                calc_result_with_margin['total'] = final_price
                
                order_details_list.append(calc_result_with_margin)
                total_price += final_price
                
            except Exception as e:
                print(f"[POST /api/orders] ❌ 아이템 계산 오류: {e}")
                raise
        
        new_order = Order(
            order_id=order_id,
            user_db_id=current_user.id,
            items=json.dumps(items, ensure_ascii=False),
            total_price=total_price,
            delivery_info=json.dumps(data.get('delivery_info', {}), ensure_ascii=False),
            order_details=json.dumps(order_details_list, ensure_ascii=False),
            status='주문접수' if ('localhost' in request.host or '127.0.0.1' in request.host) else '결제대기',  # 로컬 환경: 주문접수, 프로덕션: 결제대기
            mul_no=payment_info.get('mul_no'),
            pay_type=payment_info.get('pay_type')
        )
        
        db.session.add(new_order)
        db.session.commit()
        print(f"[POST /api/orders] ✅ 주문 저장 완료: {order_id}")
        
        # [Debug] 로컬 환경 감지 - localhost 요청이면 is_local_test 플래그 추가
        is_local = request.host.startswith('127.0.0.1') or request.host.startswith('localhost')
        print(f"[POST /api/orders] 🔧 로컬 환경 감지: {is_local} (host={request.host})")
        
        return jsonify({
            'success': True,
            'order_id': order_id,
            'customer_code': customer_code,
            'order_code': order_id,  # order_id가 이미 고객번호 포함
            'is_local_test': is_local,  # 로컬 환경이면 true
            'total_price': total_price
        })
    except Exception as e:
        db.session.rollback()
        print(f"❌ 주문 생성 에러: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'주문 생성 실패: {str(e)}'}), 500

@app.route('/api/orders/<order_id>', methods=['GET'])
def get_order_detail(order_id):
    """개별 주문 정보 조회 (로컬 환경에서는 토큰 없이도 가능)"""
    token = None
    current_user = None
    is_local = request.host.split(':')[0] in ['localhost', '127.0.0.1'] or 'localhost' in request.host
    
    if 'Authorization' in request.headers:
        try:
            auth_header = request.headers['Authorization']
            token = auth_header.split(' ')[1]
            user_id = verify_token(token)
            if user_id:
                current_user = User.query.filter_by(user_id=user_id).first()
        except:
            pass
    
    if not current_user:
        if is_local:
            test_user = User.query.filter_by(user_id='test_user').first()
            if not test_user:
                test_user = User(user_id='test_user', name='Test User', role='user')
                test_user.set_password('test')
                db.session.add(test_user)
                db.session.commit()
            current_user = test_user
        else:
            return jsonify({'success': False, 'message': '토큰이 필요합니다.'}), 401
    
    try:
        order = Order.query.filter_by(order_id=order_id, user_db_id=current_user.id).first()
        
        if not order:
            return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다'}), 404
        
        return jsonify({
            'success': True,
            'order': order.to_dict()
        })
    except Exception as e:
        print(f"❌ 주문 조회 에러: {e}")
        return jsonify({'success': False, 'message': f'주문 조회 실패: {str(e)}'}), 500

# ========== 장바구니 관련 API ==========
@app.route('/api/cart', methods=['GET'])
def get_cart():
    """장바구니 조회 (로컬 환경에서는 토큰 없이도 가능)"""
    # 토큰 확인
    token = None
    current_user = None
    is_local = request.host.split(':')[0] in ['localhost', '127.0.0.1'] or 'localhost' in request.host
    
    if 'Authorization' in request.headers:
        try:
            auth_header = request.headers['Authorization']
            token = auth_header.split(' ')[1]
            user_id = verify_token(token)
            if user_id:
                current_user = User.query.filter_by(user_id=user_id).first()
        except:
            pass
    
    # 토큰이 없으면 로컬 환경에서는 테스트 사용자 사용
    if not current_user:
        if is_local:
            print(f"[get_cart] 🔧 로컬 환경 - 테스트 사용자 사용")
            test_user = User.query.filter_by(user_id='test_user').first()
            if not test_user:
                test_user = User(user_id='test_user', name='Test User', role='user')
                test_user.set_password('test')
                db.session.add(test_user)
                db.session.commit()
            current_user = test_user
        else:
            return jsonify({'success': False, 'message': '토큰이 필요합니다.'}), 401
    
    cart_items = CartItem.query.filter_by(user_db_id=current_user.id).all()
    return jsonify({'success': True, 'cart': [item.to_dict() for item in cart_items]})

@app.route('/api/cart', methods=['POST'])
def add_to_cart():
    """장바구니에 추가 (로컬 환경에서는 토큰 없이도 가능)"""
    token = None
    current_user = None
    is_local = request.host.split(':')[0] in ['localhost', '127.0.0.1'] or 'localhost' in request.host
    
    if 'Authorization' in request.headers:
        try:
            auth_header = request.headers['Authorization']
            token = auth_header.split(' ')[1]
            user_id = verify_token(token)
            if user_id:
                current_user = User.query.filter_by(user_id=user_id).first()
        except:
            pass
    
    if not current_user:
        if is_local:
            test_user = User.query.filter_by(user_id='test_user').first()
            if not test_user:
                test_user = User(user_id='test_user', name='Test User', role='user')
                test_user.set_password('test')
                db.session.add(test_user)
                db.session.commit()
            current_user = test_user
        else:
            return jsonify({'success': False, 'message': '토큰이 필요합니다.'}), 401
    
    data = request.json
    
    cart_item = CartItem(
        user_db_id=current_user.id,
        item_data=json.dumps(data, ensure_ascii=False)
    )
    
    db.session.add(cart_item)
    db.session.commit()

    return jsonify({'success': True, 'message': '장바구니에 추가되었습니다.', 'item': cart_item.to_dict()})

@app.route('/api/cart/<int:item_id>', methods=['DELETE'])
@token_required
def delete_cart_item(current_user, item_id):
    item = CartItem.query.filter_by(user_db_id=current_user.id, id=item_id).first()
    if not item:
        return jsonify({'success': False, 'message': '항목을 찾을 수 없습니다.'}), 404
    db.session.delete(item)
    db.session.commit()
    return jsonify({'success': True, 'message': '장바구니 항목이 삭제되었습니다.'})

@app.route('/api/cart', methods=['DELETE'])
def clear_cart():
    """장바구니 비우기 (로컬 환경에서는 토큰 없이도 가능)"""
    token = None
    current_user = None
    is_local = request.host.split(':')[0] in ['localhost', '127.0.0.1'] or 'localhost' in request.host
    
    if 'Authorization' in request.headers:
        try:
            auth_header = request.headers['Authorization']
            token = auth_header.split(' ')[1]
            user_id = verify_token(token)
            if user_id:
                current_user = User.query.filter_by(user_id=user_id).first()
        except:
            pass
    
    if not current_user:
        if is_local:
            test_user = User.query.filter_by(user_id='test_user').first()
            if not test_user:
                test_user = User(user_id='test_user', name='Test User', role='user')
                test_user.set_password('test')
                db.session.add(test_user)
                db.session.commit()
            current_user = test_user
        else:
            return jsonify({'success': False, 'message': '토큰이 필요합니다.'}), 401
    
    CartItem.query.filter_by(user_db_id=current_user.id).delete()
    db.session.commit()
    
    return jsonify({'success': True, 'message': '장바구니가 비워졌습니다.'})

# ========== 공지사항 API ==========
@app.route('/api/notices', methods=['GET'])
def list_notices():
    notices = Notice.query.order_by(Notice.is_pinned.desc(), Notice.created_at.desc()).all()
    return jsonify({'success': True, 'notices': [n.to_dict() for n in notices]})

@app.route('/api/notices/<int:notice_id>', methods=['GET'])
def get_notice(notice_id):
    notice = Notice.query.get(notice_id)
    if not notice:
        return jsonify({'success': False, 'message': '공지사항을 찾을 수 없습니다.'}), 404
    return jsonify({'success': True, 'notice': notice.to_dict()})

@app.route('/api/admin/notices', methods=['POST'])
@token_required
@admin_required
def create_notice(current_user):
    data = request.json
    notice = Notice(
        title=data.get('title', ''),
        category=data.get('category', '일반공지'),
        content=data.get('content', ''),
        is_pinned=bool(data.get('is_pinned', False))
    )
    db.session.add(notice)
    db.session.commit()
    return jsonify({'success': True, 'notice': notice.to_dict()})

@app.route('/api/admin/notices/<int:notice_id>', methods=['PUT'])
@token_required
@admin_required
def update_notice(current_user, notice_id):
    notice = Notice.query.get(notice_id)
    if not notice:
        return jsonify({'success': False, 'message': '공지사항을 찾을 수 없습니다.'}), 404
    data = request.json
    notice.title = data.get('title', notice.title)
    notice.category = data.get('category', notice.category)
    notice.content = data.get('content', notice.content)
    notice.is_pinned = bool(data.get('is_pinned', notice.is_pinned))
    db.session.commit()
    return jsonify({'success': True, 'notice': notice.to_dict()})

@app.route('/api/admin/notices/<int:notice_id>', methods=['DELETE'])
@token_required
@admin_required
def delete_notice(current_user, notice_id):
    notice = Notice.query.get(notice_id)
    if not notice:
        return jsonify({'success': False, 'message': '공지사항을 찾을 수 없습니다.'}), 404
    db.session.delete(notice)
    db.session.commit()
    return jsonify({'success': True})

# ========== 비용 관리 API ==========
@app.route('/api/admin/pricing-settings', methods=['GET'])
@token_required
@admin_required
def get_pricing_settings(current_user):
    """비용 설정 조회"""
    settings = CategorySettings.query.filter_by(category='pricing').first()
    
    if settings:
        return jsonify({
            'success': True,
            'settings': json.loads(settings.settings_data)
        })
    else:
        # 기본값 반환
        return jsonify({
            'success': True,
            'settings': {
                'cover_paper_cost': 0,
                'cover_print_cost': 0,
                'cover_plate_cost': 0,
                'cover_coat_cost': 0,
                'inner_paper_cost': 0,
                'inner_print_cost': 0,
                'inner_plate_cost': 0,
                'binding_small_cost': 0,
                'binding_large_cost': 0,
                'ship_cost': 0,
                'finishing_cost': 0,
                'finishing_types': '',
                'cover_papers': '',
                'inner_papers': ''
            }
        })

@app.route('/api/admin/pricing-settings', methods=['POST'])
@token_required
@admin_required
def save_pricing_settings(current_user):
    """비용 설정 저장"""
    data = request.json
    
    settings = CategorySettings.query.filter_by(category='pricing').first()
    if not settings:
        settings = CategorySettings(category='pricing')
    
    settings.settings_data = json.dumps({
        'cover_paper_cost': float(data.get('cover_paper_cost', 0)),
        'cover_print_cost': float(data.get('cover_print_cost', 0)),
        'cover_plate_cost': float(data.get('cover_plate_cost', 0)),
        'cover_coat_cost': float(data.get('cover_coat_cost', 0)),
        'inner_paper_cost': float(data.get('inner_paper_cost', 0)),
        'inner_print_cost': float(data.get('inner_print_cost', 0)),
        'inner_plate_cost': float(data.get('inner_plate_cost', 0)),
        'binding_small_cost': float(data.get('binding_small_cost', 0)),
        'binding_large_cost': float(data.get('binding_large_cost', 0)),
        'ship_cost': float(data.get('ship_cost', 0)),
        'finishing_cost': float(data.get('finishing_cost', 0)),
        'finishing_types': data.get('finishing_types', ''),
        'cover_papers': data.get('cover_papers', ''),
        'inner_papers': data.get('inner_papers', '')
    })
    
    db.session.add(settings)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': '비용 설정이 저장되었습니다.'
    })

# ========== 팝업 공지사항 API ===============
@app.route('/api/upload-image', methods=['POST'])
@token_required
@admin_required
def upload_image(current_user):
    """이미지 파일 업로드 및 자동 최적화"""
    max_size = app.config.get('MAX_CONTENT_LENGTH')
    if max_size and request.content_length and request.content_length > max_size:
        return jsonify({'success': False, 'message': '파일 용량이 너무 큽니다. 최대 20MB까지 가능합니다.'}), 413
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '파일이 없습니다.'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '파일이 선택되지 않았습니다.'}), 400
    
    # 허용된 확장자 확인
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
    filename = file.filename.lower()
    if not any(filename.endswith('.' + ext) for ext in allowed_extensions):
        return jsonify({'success': False, 'message': '이미지 파일만 업로드 가능합니다.'}), 400
    
    # images 폴더 생성
    import os
    from werkzeug.utils import secure_filename
    from datetime import datetime
    
    upload_folder = UPLOAD_FOLDER
    os.makedirs(upload_folder, exist_ok=True)
    
    # 파일명 생성 (타임스탬프 포함)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    original_name = secure_filename(file.filename)
    name_parts = original_name.rsplit('.', 1)
    temp_filename = f"{name_parts[0]}_{timestamp}.{name_parts[1]}"
    
    # 임시로 파일 저장
    temp_filepath = os.path.join(upload_folder, temp_filename)
    file.save(temp_filepath)
    
    # 이미지 최적화 (압축 + WebP 변환)
    optimized_filepath, format_ext = optimize_image(temp_filepath)
    optimized_filename = os.path.basename(optimized_filepath)
    
    # 웹 경로 반환
    web_path = f"/images/{optimized_filename}"
    return jsonify({'success': True, 'path': web_path, 'filename': optimized_filename})

@app.route('/api/homepage-settings', methods=['GET'])
def get_homepage_settings():
    """홈페이지 설정 조회"""
    settings = HomepageSettings.query.first()
    if not settings:
        return jsonify({'success': True, 'settings': {'slides': [], 'logo': '', 'quoteImg': '', 'favicon': 'data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHZpZXdCb3g9IjAgMCA1MTIgNTEyIj48cmVjdCB3aWR0aD0iNTEyIiBoZWlnaHQ9IjUxMiIgZmlsbD0iIzI1NjNlYiIvPjx0ZXh0IHg9IjI1NiIgeT0iMzgwIiBmb250LXNpemU9IjI4MCIgZm9udC13ZWlnaHQ9ImJvbGQiIGZpbGw9IndoaXRlIiB0ZXh0LWFuY2hvcj0ibWlkZGxlIiBmb250LWZhbWlseT0iQXJpYWwsIHNhbnMtc2VyaWYiPkc8L3RleHQ+PC9zdmc+'}})
    return jsonify({'success': True, 'settings': settings.to_dict()})

@app.route('/api/homepage-settings', methods=['POST'])
@token_required
@admin_required
def save_homepage_settings(current_user):
    """홈페이지 설정 저장"""
    import os
    data = request.json
    settings = HomepageSettings.query.first()
    
    if not settings:
        settings = HomepageSettings()
        db.session.add(settings)
    
    # 기존 로고 삭제
    if 'logo' in data and data['logo']:
        if settings.logo and settings.logo != data['logo']:
            try:
                # /images/filename 형식에서 filename 추출
                old_filename = settings.logo.split('/')[-1]
                old_file = os.path.join(UPLOAD_FOLDER, old_filename)
                if os.path.exists(old_file):
                    os.remove(old_file)
                    print(f"기존 로고 삭제됨: {old_file}")
            except Exception as e:
                print(f"기존 로고 삭제 실패: {e}")
        settings.logo = data['logo']
    
    # 기존 견적 이미지 삭제
    if 'quoteImg' in data and data['quoteImg']:
        if settings.quote_img and settings.quote_img != data['quoteImg']:
            try:
                old_filename = settings.quote_img.split('/')[-1]
                old_file = os.path.join(UPLOAD_FOLDER, old_filename)
                if os.path.exists(old_file):
                    os.remove(old_file)
                    print(f"기존 견적 이미지 삭제됨: {old_file}")
            except Exception as e:
                print(f"기존 견적 이미지 삭제 실패: {e}")
        settings.quote_img = data['quoteImg']
    
    # 파비콘 처리
    if 'favicon' in data and data['favicon']:
        if settings.favicon and settings.favicon != data['favicon']:
            try:
                # /images/filename 형식에서 filename 추출
                old_filename = settings.favicon.split('/')[-1]
                old_file = os.path.join(UPLOAD_FOLDER, old_filename)
                if os.path.exists(old_file) and old_filename != 'favicon_temp.svg':
                    os.remove(old_file)
                    print(f"기존 파비콘 삭제됨: {old_file}")
            except Exception as e:
                print(f"기존 파비콘 삭제 실패: {e}")
        settings.favicon = data['favicon']
    
    # 슬라이드 이미지 삭제
    if 'slides' in data:
        try:
            # 기존 슬라이드 파싱
            if settings.slides:
                old_slides = json.loads(settings.slides)
            else:
                old_slides = []
            
            new_slides = data['slides']
            
            # 슬라이드가 문자열 배열인 경우 처리
            if old_slides and isinstance(old_slides[0], str):
                old_slide_images = set(old_slides)
                new_slide_images = set(new_slides)
            else:
                # 슬라이드가 객체 배열인 경우 처리
                old_slide_images = {s.get('image') if isinstance(s, dict) else s for s in old_slides if s}
                new_slide_images = {s.get('image') if isinstance(s, dict) else s for s in new_slides if s}
            
            # 제거된 슬라이드의 이미지 삭제
            deleted_images = old_slide_images - new_slide_images
            print(f"삭제할 이미지: {deleted_images}")
            
            for image in deleted_images:
                if image:
                    try:
                        image_filename = image.split('/')[-1]
                        image_file = os.path.join(UPLOAD_FOLDER, image_filename)
                        print(f"이미지 삭제 시도: {image_file}")
                        if os.path.exists(image_file):
                            os.remove(image_file)
                            print(f"슬라이드 이미지 삭제됨: {image_file}")
                        else:
                            print(f"파일이 존재하지 않음: {image_file}")
                    except Exception as e:
                        print(f"슬라이드 이미지 삭제 실패: {e}")
        except Exception as e:
            print(f"슬라이드 처리 중 에러: {e}")
            import traceback
            traceback.print_exc()
        
        settings.slides = json.dumps(data['slides'])
    
    db.session.commit()
    return jsonify({'success': True, 'message': '홈페이지 설정이 저장되었습니다.'})

# ==================== 카테고리 설정 API ====================
@app.route('/api/popup-notice', methods=['GET'])
def get_popup_notice():
    """활성화된 팝업 공지 조회 (전체 또는 첫 번째)"""
    popups = PopupNotice.query.filter_by(is_active=True).order_by(PopupNotice.created_at.desc()).all()
    # 팝업이 없으면 빈 배열 반환 (404 대신 200으로)
    return jsonify({
        'success': True, 
        'popup_notice': popups[0].to_dict() if popups else None,
        'popup_notices': [p.to_dict() for p in popups]
    })

@app.route('/api/admin/popup-notice', methods=['GET'])
@token_required
@admin_required
def list_popup_notices(current_user):
    """팝업 공지 목록 조회 (관리자용)"""
    popups = PopupNotice.query.order_by(PopupNotice.created_at.desc()).all()
    return jsonify({'success': True, 'popup_notices': [p.to_dict() for p in popups]})

@app.route('/api/admin/popup-notice', methods=['POST'])
@token_required
@admin_required
def create_popup_notice(current_user):
    """팝업 공지 생성"""
    data = request.json
    popup = PopupNotice(
        title=data.get('title', ''),
        image_path=data.get('image_path', ''),
        content=data.get('content', ''),
        badge=data.get('badge', ''),
        is_active=bool(data.get('is_active', True))
    )
    db.session.add(popup)
    db.session.commit()
    return jsonify({'success': True, 'popup_notice': popup.to_dict()})

@app.route('/api/admin/popup-notice/<int:popup_id>', methods=['PUT'])
@token_required
@admin_required
def update_popup_notice(current_user, popup_id):
    """팝업 공지 수정"""
    import os
    popup = PopupNotice.query.get(popup_id)
    if not popup:
        return jsonify({'success': False, 'message': '팝업 공지사항을 찾을 수 없습니다.'}), 404
    data = request.json
    
    # 기존 이미지 삭제
    if 'image_path' in data and data['image_path'] and popup.image_path != data['image_path']:
        try:
            old_filename = popup.image_path.split('/')[-1]
            old_file = os.path.join(UPLOAD_FOLDER, old_filename)
            if os.path.exists(old_file):
                os.remove(old_file)
                print(f"기존 팝업 이미지 삭제됨: {old_file}")
        except Exception as e:
            print(f"기존 팝업 이미지 삭제 실패: {e}")
    
    popup.title = data.get('title', popup.title)
    popup.image_path = data.get('image_path', popup.image_path)
    popup.content = data.get('content', popup.content)
    popup.badge = data.get('badge', popup.badge)
    popup.is_active = bool(data.get('is_active', popup.is_active))
    db.session.commit()
    return jsonify({'success': True, 'popup_notice': popup.to_dict()})

@app.route('/api/admin/popup-notice/<int:popup_id>', methods=['DELETE'])
@token_required
@admin_required
def delete_popup_notice(current_user, popup_id):
    """팝업 공지 삭제"""
    popup = PopupNotice.query.get(popup_id)
    if not popup:
        return jsonify({'success': False, 'message': '팝업 공지사항을 찾을 수 없습니다.'}), 404
    db.session.delete(popup)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/admin/images', methods=['GET'])
@token_required
@admin_required
def list_images(current_user):
    """업로드된 이미지 목록 조회"""
    import os
    try:
        if not os.path.exists(UPLOAD_FOLDER):
            return jsonify({'success': True, 'images': [], 'count': 0})
        
        files = os.listdir(UPLOAD_FOLDER)
        image_files = []
        for f in files:
            if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp')):
                filepath = os.path.join(UPLOAD_FOLDER, f)
                size = os.path.getsize(filepath)
                mtime = os.path.getmtime(filepath)
                image_files.append({
                    'filename': f,
                    'size': size,
                    'modified': mtime,
                    'url': f'/images/{f}'
                })
        
        # 수정 시간 기준 내림차순 정렬
        image_files.sort(key=lambda x: x['modified'], reverse=True)
        
        return jsonify({'success': True, 'images': image_files, 'count': len(image_files)})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/images/<filename>', methods=['DELETE'])
@token_required
@admin_required
def delete_image(current_user, filename):
    """이미지 파일 삭제"""
    import os
    try:
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'message': '파일이 존재하지 않습니다.'}), 404
        
        os.remove(filepath)
        return jsonify({'success': True, 'message': f'{filename} 삭제됨'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ========== 관리자 API ==========
@app.route('/api/admin/stats', methods=['GET'])
@token_required
@admin_required
def get_admin_stats(current_user):
    total_users = User.query.count()
    total_quotes = Quote.query.count()
    total_orders = Order.query.count()
    pending_quotes = Quote.query.filter_by(status='pending').count()
    pending_orders = Order.query.filter_by(status='결제대기').count()
    
    return jsonify({
        'success': True,
        'stats': {
            'total_users': total_users,
            'total_quotes': total_quotes,
            'total_orders': total_orders,
            'pending_quotes': pending_quotes,
            'pending_orders': pending_orders
        }
    })

@app.route('/api/admin/orders', methods=['GET'])
def list_admin_orders():
    """관리자용 주문 목록 조회 및 검색 (로컬 환경에서는 토큰 없이도 가능)"""
    # 토큰 확인 및 사용자 결정
    token = None
    current_user = None
    is_local = request.host.split(':')[0] in ['localhost', '127.0.0.1'] or 'localhost' in request.host
    
    if 'Authorization' in request.headers:
        try:
            auth_header = request.headers['Authorization']
            token = auth_header.split(' ')[1]
            user_id = verify_token(token)
            if user_id:
                current_user = User.query.filter_by(user_id=user_id).first()
        except:
            pass
    
    # 토큰이 없으면 로컬 환경에서는 테스트 관리자 사용
    if not current_user:
        if is_local:
            print(f"[list_admin_orders] 🔧 로컬 환경 - 테스트 관리자 사용")
            admin_user = User.query.filter_by(user_id='admin', role='admin').first()
            if not admin_user:
                admin_user = User(user_id='admin', name='Admin', role='admin')
                admin_user.set_password('admin')
                db.session.add(admin_user)
                db.session.commit()
            current_user = admin_user
        else:
            return jsonify({'success': False, 'message': '토큰이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
    
    search_query = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '').strip()
    
    query = Order.query.filter(Order.status != '결제대기')  # [Fix] 결제대기 상태 제외
    
    # 상태로 필터링
    if status_filter:
        query = query.filter_by(status=status_filter)
    
    # 검색어로 필터링
    if search_query:
        query = query.filter(
            ((Order.order_id.contains(search_query)) |
             (Order.user.has(User.name.contains(search_query))))
        )
    
    orders = query.order_by(Order.created_at.desc()).all()
    
    return jsonify({
        'success': True,
        'orders': [o.to_dict() for o in orders],
        'count': len(orders)
    })

@app.route('/api/admin/orders/<order_id>', methods=['GET'])
def get_admin_order(order_id):
    """관리자용 주문 상세 조회 (로컬 환경에서는 토큰 없이도 가능)"""
    # 토큰 확인 및 사용자 결정
    token = None
    current_user = None
    is_local = request.host.split(':')[0] in ['localhost', '127.0.0.1'] or 'localhost' in request.host
    
    if 'Authorization' in request.headers:
        try:
            auth_header = request.headers['Authorization']
            token = auth_header.split(' ')[1]
            user_id = verify_token(token)
            if user_id:
                current_user = User.query.filter_by(user_id=user_id).first()
        except:
            pass
    
    # 토큰이 없으면 로컬 환경에서는 테스트 관리자 사용
    if not current_user:
        if is_local:
            admin_user = User.query.filter_by(user_id='admin', role='admin').first()
            if not admin_user:
                admin_user = User(user_id='admin', name='Admin', role='admin')
                admin_user.set_password('admin')
                db.session.add(admin_user)
                db.session.commit()
            current_user = admin_user
        else:
            return jsonify({'success': False, 'message': '토큰이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
    
    order = Order.query.filter_by(order_id=order_id).first()
    if not order:
        return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
    return jsonify({'success': True, 'order': order.to_dict()})

@app.route('/api/admin/orders/<order_id>/status', methods=['PUT'])
@token_required
@admin_required
def update_order_status(current_user, order_id):
    """관리자용 주문 상태 업데이트"""
    order = Order.query.filter_by(order_id=order_id).first()
    if not order:
        return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
    
    data = request.json
    new_status = data.get('status', '').strip()
    tracking_number = data.get('tracking_number', '').strip() if data.get('tracking_number') else None
    
    # 허용된 상태
    allowed_statuses = ['결제대기', '주문접수', '제작중', '배송중', '배송완료', 'cancelled', 'refund_requested', 'refunded']
    if new_status not in allowed_statuses:
        return jsonify({'success': False, 'message': '유효하지 않은 상태입니다.'}), 400
    
    order.status = new_status
    
    # 배송중 상태일 때 송장 번호 저장
    if new_status == '배송중' and tracking_number:
        order.tracking_number = tracking_number
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f'주문 상태가 업데이트되었습니다.',
        'order': order.to_dict()
    })

@app.route('/api/admin/orders/<order_id>/shipping', methods=['PUT'])
@token_required
@admin_required
def update_shipping_info(current_user, order_id):
    """배송 정보(송장번호) 저장"""
    order = Order.query.filter_by(order_id=order_id).first()
    if not order:
        return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
    
    data = request.json
    shipping_number = data.get('shipping_number', '').strip()
    
    if not shipping_number:
        return jsonify({'success': False, 'message': '송장번호를 입력해주세요.'}), 400
    
    order.tracking_number = shipping_number
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': '송장번호가 저장되었습니다.',
        'order': order.to_dict()
    })

@app.route('/api/admin/orders/shipping/bulk-update', methods=['PUT'])
@token_required
@admin_required
def bulk_update_shipping(current_user):
    """배송 정보 일괄 업데이트"""
    data = request.json
    updates = data.get('updates', [])
    
    if not isinstance(updates, list) or len(updates) == 0:
        return jsonify({'success': False, 'message': '업데이트 데이터가 없습니다.'}), 400
    
    count = 0
    for update in updates:
        order_id = update.get('order_id', '').strip()
        shipping_number = update.get('shipping_number', '').strip()
        
        if not order_id or not shipping_number:
            continue
        
        order = Order.query.filter_by(order_id=order_id).first()
        if order:
            order.tracking_number = shipping_number
            count += 1
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f'{count}개의 송장정보가 저장되었습니다.',
        'count': count
    })

@app.route('/api/admin/orders/bulk-update-status', methods=['PUT'])
@token_required
@admin_required
def bulk_update_order_status(current_user):
    """주문 상태 일괄 변경"""
    data = request.json
    order_ids = data.get('order_ids', [])
    new_status = data.get('status', '').strip()
    
    if not isinstance(order_ids, list) or len(order_ids) == 0:
        return jsonify({'success': False, 'message': '주문이 선택되지 않았습니다.'}), 400
    
    if not new_status or new_status not in ['주문접수', '제작중', '배송중', '배송완료', '취소']:
        return jsonify({'success': False, 'message': '유효하지 않은 상태입니다.'}), 400
    
    count = 0
    for order_id in order_ids:
        order = Order.query.filter_by(order_id=order_id.strip()).first()
        if order:
            order.status = new_status
            count += 1
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f'{count}개의 주문 상태가 변경되었습니다.',
        'count': count
    })

@app.route('/api/admin/orders/bulk-delete', methods=['DELETE'])
@token_required
@admin_required
def bulk_delete_orders(current_user):
    """주문 일괄 삭제"""
    data = request.json
    order_ids = data.get('order_ids', [])
    
    if not isinstance(order_ids, list) or len(order_ids) == 0:
        return jsonify({'success': False, 'message': '주문이 선택되지 않았습니다.'}), 400
    
    count = 0
    for order_id in order_ids:
        order = Order.query.filter_by(order_id=order_id.strip()).first()
        if order:
            db.session.delete(order)
            count += 1
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f'{count}개의 주문이 삭제되었습니다.',
        'count': count
    })

@app.route('/api/orders/<order_id>/cancel', methods=['PUT', 'POST'])
@token_required
def cancel_user_order(current_user, order_id):
    """사용자가 자신의 주문을 취소 또는 삭제"""
    order = Order.query.filter_by(order_id=order_id).first()
    if not order:
        return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
    
    # 현재 사용자의 주문인지 확인
    if order.user_db_id != current_user.id:
        return jsonify({'success': False, 'message': '다른 사용자의 주문은 취소할 수 없습니다.'}), 403
    
    # [Fix] 결제 미완료(pending) 상태면 삭제, 아니면 취소 상태로 변경
    if order.status == '결제대기':
        # pending 상태면 DELETE처럼 처리
        db.session.delete(order)
        db.session.commit()
        print(f"[cancel_user_order] ✅ 미결제 주문 삭제 완료: {order_id}")
        return jsonify({
            'success': True,
            'message': '결제 미완료 주문이 삭제되었습니다.',
            'order_id': order_id
        })
    elif order.status == 'cancelled':
        return jsonify({'success': False, 'message': '이미 취소된 주문입니다.'}), 400
    else:
        # 다른 상태면 취소 처리
        order.status = 'cancelled'
        db.session.commit()
        print(f"[cancel_user_order] ✅ 주문 취소 완료: {order_id}")
        return jsonify({
            'success': True,
            'message': '주문이 취소되었습니다.',
            'order': order.to_dict()
        })

@app.route('/api/orders/<order_id>', methods=['PUT'])
@token_required
def update_order_payment(current_user, order_id):
    """주문의 결제 정보 업데이트 (mul_no, pay_type 등)"""
    print(f"[PUT /api/orders/{order_id}] 요청 사용자: {current_user.id}")
    
    order = Order.query.filter_by(order_id=order_id).first()
    if not order:
        print(f"[PUT] 주문을 찾을 수 없음: {order_id}")
        return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
    
    # 현재 사용자의 주문인지 확인
    if order.user_db_id != current_user.id:
        print(f"[PUT] 권한 없음: 요청자={current_user.id}, 주문자={order.user_db_id}")
        return jsonify({'success': False, 'message': '다른 사용자의 주문은 수정할 수 없습니다.'}), 403
    
    data = request.json
    
    # mul_no, pay_type 업데이트
    if 'mul_no' in data:
        order.mul_no = data['mul_no']
        print(f"[PUT] mul_no 업데이트: {data['mul_no']}")
    
    if 'pay_type' in data:
        order.pay_type = data['pay_type']
        print(f"[PUT] pay_type 업데이트: {data['pay_type']}")
    
    db.session.commit()
    print(f"[PUT] ✅ 주문 정보 업데이트 완료: {order_id}")
    
    return jsonify({
        'success': True,
        'message': '주문이 업데이트되었습니다.',
        'order_id': order_id
    })

# [Fix] PATCH: 결제 완료 후 상태 업데이트
@app.route('/api/orders/<order_id>', methods=['PATCH'])
@token_required
def update_order_status_on_payment(current_user, order_id):
    """주문 상태 업데이트 (pending → completed) - 결제 완료 시"""
    print(f"[PATCH /api/orders/{order_id}] 요청 사용자: {current_user.id}")
    
    order = Order.query.filter_by(order_id=order_id).first()
    if not order:
        print(f"[PATCH] 주문을 찾을 수 없음: {order_id}")
        return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
    
    # 현재 사용자의 주문인지 확인
    if order.user_db_id != current_user.id:
        print(f"[PATCH] 권한 없음: 요청자={current_user.id}, 주문자={order.user_db_id}")
        return jsonify({'success': False, 'message': '다른 사용자의 주문은 수정할 수 없습니다.'}), 403
    
    data = request.json
    
    # status 업데이트 (pending → completed)
    if 'status' in data:
        order.status = data['status']
        print(f"[PATCH] status 업데이트: {data['status']}")
    
    db.session.commit()
    print(f"[PATCH] ✅ 주문 상태 업데이트 완료: {order_id} → {order.status}")
    
    return jsonify({
        'success': True,
        'message': f'주문 상태가 {order.status}로 업데이트되었습니다.',
        'order_id': order_id,
        'status': order.status
    })

@app.route('/api/orders/<order_id>', methods=['DELETE'])
@token_required
def delete_user_order(current_user, order_id):
    """
    사용자가 미완료 주문 또는 취소된 주문을 삭제
    - 결제 미완료 (mul_no 없음): 바로 삭제 가능
    - 취소된 주문: 삭제 가능
    """
    print(f"[DELETE /api/orders/{order_id}] 요청 사용자: {current_user.id}")
    
    order = Order.query.filter_by(order_id=order_id).first()
    if not order:
        print(f"[DELETE] 주문을 찾을 수 없음: {order_id}")
        return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
    
    print(f"[DELETE] 주문 정보 - status: {order.status}, mul_no: {order.mul_no}, user_db_id: {order.user_db_id}, current_user.id: {current_user.id}")
    
    # 현재 사용자의 주문인지 확인
    if order.user_db_id != current_user.id:
        print(f"[DELETE] 권한 없음: 요청자={current_user.id}, 주문자={order.user_db_id}")
        return jsonify({'success': False, 'message': '다른 사용자의 주문은 삭제할 수 없습니다.'}), 403
    
    # 삭제 가능한 경우:
    # 1. 결제 미완료 상태 (mul_no가 없고 pending 상태)
    # 2. 취소된 주문 (status='cancelled')
    is_unpaid = not order.mul_no and order.status == '결제대기'
    is_cancelled = order.status == 'cancelled'
    
    print(f"[DELETE] 삭제 가능 여부 - is_unpaid: {is_unpaid}, is_cancelled: {is_cancelled}")
    
    if not (is_unpaid or is_cancelled):
        print(f"[DELETE] 삭제 불가능: status={order.status}, mul_no={order.mul_no}")
        return jsonify({
            'success': False, 
            'message': '결제 미완료 또는 취소된 주문만 삭제할 수 있습니다.',
            'order_status': order.status,
            'mul_no': order.mul_no,
            'is_unpaid': is_unpaid,
            'is_cancelled': is_cancelled
        }), 400
    
    db.session.delete(order)
    db.session.commit()
    print(f"[DELETE] ✅ 주문 삭제 완료: {order_id}")
    
    return jsonify({
        'success': True,
        'message': '주문이 삭제되었습니다.'
    })

@app.route('/api/orders/<order_id>/refund', methods=['POST'])
@token_required
def request_refund(current_user, order_id):
    """사용자가 환불을 요청"""
    order = Order.query.filter_by(order_id=order_id).first()
    if not order:
        return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
    
    # 현재 사용자의 주문인지 확인
    if order.user_db_id != current_user.id:
        return jsonify({'success': False, 'message': '다른 사용자의 주문은 환불 요청할 수 없습니다.'}), 403
    
    # 환불 가능한 상태 확인 (주문접수, 제작중만 가능)
    if order.status not in ['주문접수', '제작중']:
        return jsonify({'success': False, 'message': '이 상태에서는 환불을 요청할 수 없습니다.'}), 400
    
    # 이미 환불 요청했거나 환불 완료된 경우
    if order.status in ['refund_requested', 'refunded']:
        return jsonify({'success': False, 'message': '이미 환불이 요청되었거나 완료되었습니다.'}), 400
    
    order.status = 'refund_requested'
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': '환불 요청이 접수되었습니다.',
        'order': order.to_dict()
    })

@app.route('/api/admin/orders/<order_id>/refund/approve', methods=['PUT'])
@token_required
@admin_required
def approve_refund(current_user, order_id):
    """관리자가 환불을 승인 - PayApp 결제 취소"""
    print(f"🔄 환불 승인 시작: {order_id}")
    order = Order.query.filter_by(order_id=order_id).first()
    if not order:
        print(f"❌ 주문을 찾을 수 없음: {order_id}")
        return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
    
    print(f"📋 주문 상태: {order.status}, mul_no: {order.mul_no}")
    
    if order.status != 'refund_requested':
        print(f"❌ 환불 요청 상태 아님: {order.status}")
        return jsonify({'success': False, 'message': '환불 요청 상태의 주문만 승인할 수 있습니다.'}), 400
    
    try:
        # PayApp 결제 취소 처리 (mul_no가 있는 경우만)
        if order.mul_no:
            print(f"🔗 PayApp 취소 API 호출 시작: mul_no={order.mul_no}")
            payapp_response = _cancel_payapp_payment(
                mul_no=order.mul_no,
                total_price=order.total_price,
                cancel_memo='테스트 환불 처리'
            )
            
            if not payapp_response.get('success'):
                return jsonify({
                    'success': False, 
                    'message': f'PayApp 취소 실패: {payapp_response.get("error", "알 수 없는 오류")}'
                }), 500
            
            print(f"✅ PayApp 취소 성공: {order_id}")
        else:
            # mul_no가 없는 경우 수동 환불로 처리
            print(f"📋 수동 환불 처리: {order_id} (결제번호 없음)")
        
        # 주문 상태 업데이트
        order.status = 'refunded'
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '환불이 처리되었습니다.',
            'order': order.to_dict()
        })
    
    except Exception as e:
        print(f"❌ 환불 처리 중 오류: {e}")
        return jsonify({
            'success': False,
            'message': f'환불 처리 중 오류 발생: {str(e)}'
        }), 500

@app.route('/api/admin/orders/<order_id>/refund/reject', methods=['PUT'])
@token_required
@admin_required
def reject_refund(current_user, order_id):
    """관리자가 환불을 거절"""
    data = request.json or {}
    reason = data.get('reason', '관리자 판단')
    
    order = Order.query.filter_by(order_id=order_id).first()
    if not order:
        return jsonify({'success': False, 'message': '주문을 찾을 수 없습니다.'}), 404
    
    if order.status != 'refund_requested':
        return jsonify({'success': False, 'message': '환불 요청 상태의 주문만 거절할 수 있습니다.'}), 400
    
    # 환불 거절 시 이전 상태로 복원 (결제대기)
    order.status = '결제대기'
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f'환불이 거절되었습니다. 사유: {reason}',
        'order': order.to_dict()
    })

# ==================== 개인결제 링크 API ====================

@app.route('/api/payment-links', methods=['GET'])
@token_required
@admin_required
def get_payment_links(current_user):
    """관리자: 생성한 개인결제 링크 목록 조회"""
    try:
        links = PaymentLink.query.order_by(PaymentLink.created_at.desc()).all()
        return jsonify({
            'success': True,
            'links': [link.to_dict() for link in links]
        })
    except Exception as e:
        print(f"❌ 링크 목록 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/payment-links', methods=['POST'])
@token_required
@admin_required
def create_payment_link(current_user):
    """관리자: 개인결제 링크 생성"""
    try:
        data = request.json or {}
        product_name = data.get('product_name', '').strip()
        price = data.get('price', 0)
        customer_name = data.get('customer_name', '').strip()
        customer_phone = data.get('customer_phone', '').strip()
        memo = data.get('memo', '').strip()

        if not product_name:
            return jsonify({'success': False, 'message': '상품명을 입력해주세요.'}), 400
        
        if not price or price < 1000:
            return jsonify({'success': False, 'message': '결제금액은 1,000원 이상이어야 합니다.'}), 400

        # 고유 링크 코드 생성 (8자리 영문+숫자)
        link_code = secrets.token_urlsafe(6).upper()[:8]
        while PaymentLink.query.filter_by(link_code=link_code).first():
            link_code = secrets.token_urlsafe(6).upper()[:8]

        new_link = PaymentLink(
            link_code=link_code,
            product_name=product_name,
            price=float(price),
            customer_name=customer_name or None,
            customer_phone=customer_phone or None,
            memo=memo or None,
            created_by=current_user.user_id
        )

        db.session.add(new_link)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': '개인결제 링크가 생성되었습니다.',
            'link': new_link.to_dict()
        })

    except Exception as e:
        db.session.rollback()
        print(f"❌ 링크 생성 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/payment-links/<link_code>', methods=['GET'])
def get_payment_link_by_code(link_code):
    """고객: 링크 코드로 결제 정보 조회 (인증 불필요)"""
    try:
        link = PaymentLink.query.filter_by(link_code=link_code).first()
        if not link:
            return jsonify({'success': False, 'message': '유효하지 않은 링크입니다.'}), 404
        
        if link.is_used:
            return jsonify({'success': False, 'message': '이미 사용된 링크입니다.'}), 400

        return jsonify({
            'success': True,
            'link': link.to_dict()
        })
    except Exception as e:
        print(f"❌ 링크 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/payment-links/<link_code>/use', methods=['POST'])
def use_payment_link(link_code):
    """고객: 개인결제 링크로 결제 완료 후 주문 생성 (인증 불필요)"""
    try:
        data = request.json or {}
        order_id = data.get('order_id')
        
        if not order_id:
            return jsonify({'success': False, 'message': '주문번호가 필요합니다.'}), 400

        link = PaymentLink.query.filter_by(link_code=link_code).first()
        if not link:
            return jsonify({'success': False, 'message': '유효하지 않은 링크입니다.'}), 404

        if link.is_used:
            return jsonify({'success': False, 'message': '이미 사용된 링크입니다.'}), 400

        # 링크 사용 처리
        link.is_used = True
        link.order_id = order_id
        link.used_at = datetime.utcnow()
        db.session.commit()

        return jsonify({
            'success': True,
            'message': '결제가 완료되었습니다.',
            'link': link.to_dict()
        })

    except Exception as e:
        db.session.rollback()
        print(f"❌ 링크 사용 처리 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/payment-links/<int:link_id>', methods=['DELETE'])
@token_required
@admin_required
def delete_payment_link(current_user, link_id):
    """관리자: 링크 삭제 (미사용만 가능)"""
    try:
        link = PaymentLink.query.get(link_id)
        if not link:
            return jsonify({'success': False, 'message': '링크를 찾을 수 없습니다.'}), 404

        if link.is_used:
            return jsonify({'success': False, 'message': '이미 사용된 링크는 삭제할 수 없습니다.'}), 400

        db.session.delete(link)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': '링크가 삭제되었습니다.'
        })

    except Exception as e:
        db.session.rollback()
        print(f"❌ 링크 삭제 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# [Migration] 상태값 영문->한글로 변환 (앱 시작 시 자동 실행)
def migrate_status_to_korean():
    """기존 데이터베이스의 영문 상태값을 한글로 변환"""
    try:
        status_mapping = {
            'pending': '결제대기',
            'completed': '주문접수',
            'preparing': '제작중',
            'shipping': '배송중',
            'delivered': '배송완료'
        }
        
        total_updated = 0
        for eng_status, kor_status in status_mapping.items():
            count = Order.query.filter_by(status=eng_status).count()
            if count > 0:
                Order.query.filter_by(status=eng_status).update({'status': kor_status})
                db.session.commit()
                print(f"[Migration] '{eng_status}' → '{kor_status}': {count}개 업데이트")
                total_updated += count
        
        if total_updated > 0:
            print(f"[Migration] 총 {total_updated}개의 주문 상태를 한글로 변환 완료")
        else:
            print(f"[Migration] 변환할 주문이 없습니다 (이미 한글로 변환됨)")
    except Exception as e:
        print(f"[Migration] 마이그레이션 중 에러: {e}")

# ========== 엑셀 다운로드/업로드 기능 ==========
@app.route('/api/admin/orders/export-excel', methods=['GET'])
@token_required
def export_orders_excel(current_user):
    """주문 목록을 엑셀로 다운로드 (관리자만)"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': '관리자만 이용 가능합니다.'}), 403
    
    try:
        # 모든 주문 조회
        orders = Order.query.order_by(Order.order_id.desc()).all()
        
        # 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "주문목록"
        
        # 스타일 설정
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 헤더 행
        headers = ['주문ID', '고객명', '전화', '이메일', '주문상태', '상품명', '수량', '가격', '배송사', '송장번호', '주문날짜']
        ws.append(headers)
        
        # 헤더 스타일 적용
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # 데이터 행
        for order in orders:
            ws.append([
                order.order_id,
                order.name,
                order.phone,
                order.email,
                order.status,
                order.product_name,
                order.quantity,
                order.total_price,
                '로젠택배',  # 기본값
                order.tracking_number or '',
                order.order_date.strftime('%Y-%m-%d %H:%M') if order.order_date else ''
            ])
        
        # 컬럼 너비 자동 조정
        column_widths = [12, 12, 12, 18, 12, 20, 8, 10, 10, 15, 16]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        
        # 데이터 셀 스타일 적용
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
            for cell in row:
                cell.border = border
                cell.alignment = center_align
        
        # 바이트 스트림으로 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 파일명
        filename = f"주문목록_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_from_directory(
            directory=os.path.dirname(output.getvalue()),
            path=filename,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        print(f"[export_excel] ❌ 엑셀 다운로드 중 에러: {e}")
        return jsonify({'success': False, 'message': f'엑셀 다운로드 실패: {str(e)}'}), 500

@app.route('/api/admin/orders/import-excel', methods=['POST'])
@token_required
def import_orders_excel(current_user):
    """엑셀 파일로 주문 일괄 업로드 (송장번호 등)"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': '관리자만 이용 가능합니다.'}), 403
    
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '파일이 없습니다.'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '파일을 선택해주세요.'}), 400
        
        # 엑셀 파일 읽기
        from openpyxl import load_workbook
        
        file_stream = BytesIO(file.read())
        wb = load_workbook(file_stream)
        ws = wb.active
        
        updated_count = 0
        error_list = []
        
        # 첫 행(헤더) 제외하고 데이터 처리
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                order_id = row[0]  # 주문ID
                tracking_number = row[9]  # 송장번호
                
                if not order_id:
                    continue
                
                # 주문 조회
                order = Order.query.filter_by(order_id=str(order_id)).first()
                if not order:
                    error_list.append(f"행 {row_idx}: 주문 {order_id}을 찾을 수 없습니다.")
                    continue
                
                # 송장번호가 입력된 경우
                if tracking_number and tracking_number.strip():
                    order.tracking_number = str(tracking_number).strip()
                    # 상태를 배송중으로 자동 변경
                    if order.status != '배송중':
                        order.status = '배송중'
                    updated_count += 1
            
            except Exception as e:
                error_list.append(f"행 {row_idx}: 처리 중 오류 - {str(e)}")
        
        db.session.commit()
        
        message = f"✅ {updated_count}개 주문이 업로드되었습니다."
        if error_list:
            message += f"\n❌ {len(error_list)}개 항목 오류:\n" + "\n".join(error_list[:5])
        
        return jsonify({
            'success': True,
            'message': message,
            'updated_count': updated_count,
            'error_count': len(error_list)
        })
    
    except Exception as e:
        print(f"[import_excel] ❌ 엑셀 업로드 중 에러: {e}")
        return jsonify({'success': False, 'message': f'엑셀 업로드 실패: {str(e)}'}), 500

@app.route('/api/calculate-quote', methods=['POST'])
def calculate_quote():
    """견적 계산 (백엔드)"""
    try:
        data = request.get_json()
        category = data.get('category')  # 'indigo', 'digital', 'offset', 'flyer_small', 'flyer_large'
        qty = data.get('qty', 0)
        margin_override = data.get('margin')  # 프론트에서 마진 오버라이드
        specs = data.get('specs', {})  # 종이 종류, 평량 등
        member_type = data.get('member_type', 'general')  # 회원 유형
        
        if not category or qty <= 0:
            return jsonify({'success': False, 'message': '카테고리와 수량을 확인해주세요.'}), 400
        
        # 마진: 프론트 오버라이드 > DB 조회 > 기본값 0
        if margin_override is not None:
            margin = margin_override
        else:
            # DB에서 카테고리별 마진 조회
            category_map = {
                'flyer_small': 'flyer_small',
                'flyer_large': 'flyer_large',
                'indigo': 'indigo',
                'digital': 'digital',
                'offset': 'offset'
            }
            db_category = category_map.get(category, category)
            margin = get_margin_from_db(db_category, member_type)
        
        # 카테고리별 계산
        if category.startswith('flyer'):
            result = calculate_flyer_backend(category, qty, margin, {}, specs)
        else:  # indigo, digital, offset
            result = calculate_book_backend(category, qty, margin, {}, specs)
        
        return jsonify({
            'success': True,
            'data': result
        })
    except Exception as e:
        print(f"[ERROR] 견적 계산 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/category-costs', methods=['GET'])
def get_category_costs():
    """카테고리별 비용 조회"""
    print("[DEBUG] get_category_costs() 호출됨")
    try:
        categories = ['indigo', 'digital', 'offset', 'flyer_small', 'flyer_large']
        result = {}
        
        for category in categories:
            # 마진율 조회
            cat_setting = CategorySettings.query.filter_by(category=category).first()
            margin = 0
            if cat_setting:
                settings = json.loads(cat_setting.settings_data)
                margin = settings.get('margin', 0)
            
            # 인쇄비 조회
            print_costs = PrintCost.query.filter_by(category=category).all()
            print_cost_dict = {}
            for pc in print_costs:
                print_cost_dict[pc.print_type] = pc.cost_per_click
            
            # 판비 조회
            plate_costs = PlateCost.query.filter_by(category=category).all()
            plate_cost_dict = {}
            for pc in plate_costs:
                plate_cost_dict[pc.plate_unit_name] = {
                    'base_cost': pc.base_cost,
                    'per_color': pc.per_color
                }
            
            result[category] = {
                'margin': margin,
                'print_costs': print_cost_dict,
                'plate_costs': plate_cost_dict
            }
        
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        print(f"[ERROR] 카테고리 비용 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/category-costs/<category>', methods=['POST'])
def update_category_costs(category):
    """카테고리별 비용 저장"""
    try:
        data = request.get_json()
        
        # 마진율 저장 (CategorySettings)
        margin = data.get('margin', 0)
        cat_setting = CategorySettings.query.filter_by(category=category).first()
        
        if cat_setting:
            settings = json.loads(cat_setting.settings_data)
            settings['margin'] = margin
            cat_setting.settings_data = json.dumps(settings)
        else:
            cat_setting = CategorySettings(
                category=category,
                settings_data=json.dumps({'margin': margin})
            )
            db.session.add(cat_setting)
        
        db.session.commit()
        return jsonify({'success': True, 'message': f'{category} 비용이 저장되었습니다.'})
    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] 카테고리 비용 저장 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

def get_margin_from_db(category, is_member_type='general'):
    """DB에서 카테고리/회원유형별 마진율 조회"""
    margin_setting = MarginSetting.query.filter_by(
        category=category,
        member_type=is_member_type
    ).first()
    if margin_setting:
        return margin_setting.margin_rate
    
    # 기본 마진 0%
    return 0

def get_print_cost_from_db(print_type, default_value):
    """DB에서 인쇄비 조회 (indigo_color, indigo_mono, digital, offset)"""
    cost = AdditionalCost.query.filter_by(cost_name=f'print_{print_type}').first()
    if cost:
        print(f"[DB 조회] 인쇄비({print_type}): {int(cost.cost)}원 (DB에서)")
        return int(cost.cost)
    print(f"[폴백] 인쇄비({print_type}): {default_value}원 (기본값)")
    return default_value

def get_paper_price_from_db(paper_type, gram, is_4x6=False):
    """DB에서 종이 가격 조회"""
    paper = PaperPrice.query.filter_by(paper_type=paper_type, gram=gram).first()
    if not paper:
        return None
    return paper.sheet_4x6_price if is_4x6 else paper.kook_price

def get_shipping_cost_from_db(category='flyer_large'):
    """DB에서 카테고리별 배송비 조회"""
    cost_name = f'shipping_{category}'
    cost = AdditionalCost.query.filter_by(cost_name=cost_name).first()
    if cost:
        print(f"[DB 조회] 배송비({category}): {int(cost.cost)}원 (DB에서)")
        return int(cost.cost)
    print(f"[폴백] 배송비({category}): 3000원 (기본값)")
    return 3000

# 박스 정의 (mm, kg)
BOX_A4 = {'name': 'A4박스', 'w': 315, 'l': 220, 'h': 270, 'maxKg': 20}
BOX_A3 = {'name': 'A3박스', 'w': 450, 'l': 305, 'h': 210, 'maxKg': 20}

# 두께 데이터 (gram -> mm)
THICKNESS_DB = {
    80: 0.09, 100: 0.105, 120: 0.13, 150: 0.16,
    180: 0.19, 200: 0.21, 250: 0.26, 300: 0.31
}

def get_thickness_by_gram(gram):
    """그램별 두께 조회"""
    return THICKNESS_DB.get(gram, 0.1)

def calculate_shipping_cost(qty, size, is_flyer, cv_gram, in_gram, inner_pages, category='flyer_large'):
    """
    HTML의 calculateShipping과 동일한 로직
    배송비 = 필요 박스 수 × 박스 가격 (DB에서 조회)
    """
    # 사이즈별 mm
    size_map = {
        'A4': (210, 297),
        'A5': (148, 210),
        'B5': (182, 257)
    }
    w_mm, h_mm = size_map.get(size, (210, 297))
    area = (w_mm / 1000) * (h_mm / 1000)
    
    # 개당 무게 계산
    single_weight = 0
    single_thick = 0
    
    if is_flyer:
        single_weight = area * in_gram
        single_thick = get_thickness_by_gram(in_gram)
    else:
        # 책자: 내지 + 표지
        in_sheets = math.ceil(inner_pages / 2)
        single_weight += area * in_gram * in_sheets
        single_thick += get_thickness_by_gram(in_gram) * in_sheets
        
        # 표지 (보통 4페이지 = 2장)
        cv_sheets = 2
        single_weight += (area * 2) * cv_gram * cv_sheets
        single_thick += get_thickness_by_gram(cv_gram) * cv_sheets
    
    # 박스 선택
    box = BOX_A3 if is_flyer else BOX_A4
    
    # 박스당 수량 계산
    books_per_layer = 1
    if not is_flyer and size == 'A5':
        books_per_layer = 2
    elif is_flyer and size == 'A4':
        books_per_layer = 2
    elif is_flyer and size == 'A5':
        books_per_layer = 4
    
    books_per_stack = int(box['h'] / single_thick) if single_thick > 0 else 1
    max_books_by_vol = books_per_stack * books_per_layer
    if max_books_by_vol < 1:
        max_books_by_vol = 1
    
    max_books_by_weight = int((box['maxKg'] * 1000) / single_weight) if single_weight > 0 else 1
    count_per_box = min(max_books_by_vol, max_books_by_weight)
    
    total_boxes = math.ceil(qty / count_per_box)
    
    # DB에서 배송비 조회
    box_price = get_shipping_cost_from_db(category)
    total_ship_cost = total_boxes * box_price
    
    return {
        'cost': total_ship_cost,
        'boxes': total_boxes,
        'box_name': box['name']
    }

def calculate_flyer_backend(category, qty, margin, costs, specs):
    """백엔드 전단지 계산 (소량/대량 전단)"""
    try:
        # specs가 string이면 기본값으로 처리 (호환성)
        if isinstance(specs, str):
            print(f"[WARN] specs이 string입니다 (호환성 모드): {specs}")
            specs = {}
        
        size = specs.get('size', 'A4')
        in_type = specs.get('inner_type', '모조지')
        in_gram = specs.get('inner_gram', '80')
        in_color = specs.get('inner_color', 'color')
        is_double = specs.get('is_double', True)
        
        print(f"\n[FLYER 계산] category={category}, qty={qty}, size={size}, margin={margin}")
        
        # 종이 가격 조회
        in_price = get_paper_price_from_db(in_type, int(in_gram))
        print(f"[DB 조회] 종이가격({in_type} {in_gram}g): {in_price}원 (DB에서)")
        if in_price is None:
            in_price = YEON_PRICE_111.get(in_type, {}).get(str(in_gram), {}).get('kook', 50000)
            print(f"[폴백] 종이가격 사용: {in_price}원 (기본값)")
        
        p_paper = 0
        p_print = 0
        p_plate = 0
        
        if category == 'flyer_small':
            # [소량 전단 - 인디고]
            yield_small = {'A4': 2, 'A5': 4, 'B5': 2}.get(size, 2)
            sheets_a3 = int((qty + yield_small - 1) / yield_small)
            price_a3 = in_price / 2000
            p_paper = int(sheets_a3 * price_a3)
            
            # 인쇄비 (DB에서 조회)
            click_color = get_print_cost_from_db('indigo_color', 200)
            click_mono = get_print_cost_from_db('indigo_mono', 40)
            click_unit = click_color if in_color == 'color' else click_mono
            final_click = click_unit if is_double else (click_unit / 2)
            p_print = int(sheets_a3 * final_click)
            
            # 배송비 계산
            ship_info = calculate_shipping_cost(qty, size, True, 0, int(in_gram), 0, 'flyer_small')
            ship_cost = ship_info['cost']
        
        else:  # flyer_large
            # [대량 전단 - 옵셋]
            yield_large = {'A4': 8, 'A5': 16, 'B5': 8}.get(size, 8)
            plates = 4 if in_color == 'color' else 1
            plates = plates * 2 if is_double else plates
            
            # 판비 (DB에서 조회)
            plate_cost = AdditionalCost.query.filter_by(cost_name='plate_flyer_large').first()
            plate_price = int(plate_cost.cost) if plate_cost else 11000
            p_plate = int(plates * plate_price)
            
            # 종이비 + 인쇄비
            sheets_full = int((qty / yield_large) + 100)
            yeon = sheets_full / 500
            p_paper = int(yeon * in_price)
            
            # 인쇄비 (색당, DB에서 조회)
            offset_price = get_print_cost_from_db('offset', 8000)
            p_print = int(yeon * plates * offset_price)
            
            # 배송비 계산
            ship_info = calculate_shipping_cost(qty, size, True, 0, int(in_gram), 0, 'flyer_large')
            ship_cost = ship_info['cost']
        
        # 배송비 제외 비용으로 계산
        cost_without_shipping = p_paper + p_print + p_plate
        
        # 마진 적용 (배송비 제외)
        supply_cost = int(cost_without_shipping * (1 + margin / 100))
        
        # 부가세는 공금가 + 배송비에 적용
        vat = int((supply_cost + ship_cost) * 0.1)
        total_price = int((supply_cost + ship_cost + vat) / 10) * 10
        
        print(f"\n[FLYER 최종결과]")
        print(f"  - 종이비: {p_paper}원")
        print(f"  - 인쇄비: {p_print}원")
        print(f"  - 판비: {p_plate}원")
        print(f"  - 배송비: {ship_cost}원")
        print(f"  - 공금가: {supply_cost}원 (마진 {margin}% 적용)")
        print(f"  - 부가세: {vat}원")
        print(f"  - 최종가: {total_price}원")
        
        return {
            'category': category,
            'qty': qty,
            'supply_cost': supply_cost,
            'vat': vat,
            'shipping': ship_cost,
            'total': total_price,
            'breakdown': {
                'paper': p_paper,
                'print': p_print,
                'plate': p_plate,
                'shipping': ship_cost
            }
        }
    except Exception as e:
        print(f"❌ 전단지 계산 오류: {e}")
        raise

def calculate_book_backend(category, qty, margin, costs, specs):
    """백엔드 책자 계산 (소량/대량 책자)"""
    try:
        # specs가 string이면 기본값으로 처리 (호환성)
        if isinstance(specs, str):
            print(f"[WARN] specs이 string입니다 (호환성 모드): {specs}")
            specs = {}
        
        size = specs.get('size', 'A4')
        inner_pages = specs.get('inner_pages', 0)
        cv_type = specs.get('cover_type', '모조지')
        cv_gram = specs.get('cover_gram', '100')
        in_type = specs.get('inner_type', '모조지')
        in_gram = specs.get('inner_gram', '80')
        bind_type = specs.get('bind_type', 'perfect')
        cv_color = specs.get('cover_color', 'color')
        in_color = specs.get('inner_color', 'color')
        coating = specs.get('coating', '0')
        cover_page = specs.get('cover_page', 4)  # 2 또는 4
        
        print(f"\n[BOOK 계산] category={category}, qty={qty}, size={size}, inner_pages={inner_pages}, margin={margin}")
        print(f"  표지: {cv_type} {cv_gram}g {cv_color}")
        print(f"  내지: {in_type} {in_gram}g {in_color}")
        
        # 종이 가격 조회
        cv_price = get_paper_price_from_db(cv_type, int(cv_gram))
        print(f"[DB 조회] 표지 종이가격({cv_type} {cv_gram}g): {cv_price}원")
        if cv_price is None:
            cv_price = YEON_PRICE_111.get(cv_type, {}).get(str(cv_gram), {}).get('kook', 50000)
            print(f"[폴백] 표지 종이가격 사용: {cv_price}원 (기본값)")
        
        in_price = get_paper_price_from_db(in_type, int(in_gram))
        print(f"[DB 조회] 내지 종이가격({in_type} {in_gram}g): {in_price}원")
        if in_price is None:
            in_price = YEON_PRICE_111.get(in_type, {}).get(str(in_gram), {}).get('kook', 50000)
        
        cv_p = 0
        cv_pr = 0
        cv_pl = 0
        cv_c = 0
        in_p = 0
        in_pr = 0
        in_pl = 0
        bind = 0
        
        # 배송비 카테고리 설정
        ship_category = 'offset' if category == 'offset' else ('digital' if category == 'digital' else 'indigo')
        
        # 배송비 계산 (책자는 A4박스)
        ship_info = calculate_shipping_cost(qty, size, False, int(cv_gram), int(in_gram), inner_pages, ship_category)
        ship_cost = ship_info['cost']
        
        if category == 'offset':
            # [대량 책자 - 옵셋]
            pages_per_form = 32 if size == 'A5' else 16
            
            # 표지 계산
            if coating != '0':
                cv_c = 45000 if qty <= 500 else (80000 if qty <= 1000 else 120000)
            
            covers_per_sheet = 4 if size == 'A5' else 2
            cv_sheets_full = int((qty / covers_per_sheet) + 130)
            cv_yeon = cv_sheets_full / 1000
            cv_p = int(cv_yeon * cv_price)
            
            cv_plates = 8 if cover_page == 4 else 4
            cv_pl = int(cv_plates * 8000)
            cv_pr = int(max(1, cv_yeon) * cv_plates * 8000)
            
            # 내지 계산
            daesu = math.ceil((inner_pages / pages_per_form) * 2) / 2
            in_sheets_total = (daesu * qty) + (daesu * 130)
            yeon = in_sheets_total / 500
            
            in_plate_count = (4 if in_color == 'color' else 1) * 2
            in_pl = int(daesu * in_plate_count * 11000)
            in_p = int(yeon * in_price)
            offset_price = get_print_cost_from_db('offset', 8000)
            in_pr = int(max(1, yeon) * in_plate_count * offset_price)
            
            # 제본비
            if bind_type == 'perfect':
                div = 4000 if size == 'A4' else 8000
                bind_r = int((inner_pages / 2 * qty) / div)
                bind = 120000 + (max(0, bind_r - 6) * 20000)
            else:
                bind = 50000 + (qty * 300)
        
        else:
            # [소량 책자 - 인디고/디지털]
            cv_sheet = cv_price / 2000
            cv_p = int(qty * cv_sheet)
            
            # 표지 인쇄비 (DB에서 조회)
            if category == 'digital':
                c_click = get_print_cost_from_db('digital', 200)
            else:
                click_color = get_print_cost_from_db('indigo_color', 200)
                click_mono = get_print_cost_from_db('indigo_mono', 40)
                c_click = click_color if cv_color == 'color' else click_mono
            final_click = c_click if cover_page == 4 else (c_click / 2)
            cv_pr = int(qty * final_click)
            
            if coating != '0':
                cv_c = qty * 300
            
            # 내지
            in_sheet = in_price / 2000
            factor = 8 if size == 'A5' else 4
            sheets = math.ceil(inner_pages / factor) * qty
            in_p = int(sheets * in_sheet)
            
            # 내지 인쇄비 (DB에서 조회)
            if category == 'digital':
                i_click = get_print_cost_from_db('digital', 20)
            else:
                click_color = get_print_cost_from_db('indigo_color', 200)
                click_mono = get_print_cost_from_db('indigo_mono', 40)
                i_click = click_color if in_color == 'color' else click_mono
            in_pr = int(sheets * i_click)
            
            # 제본비
            bind = qty * (200 if bind_type == 'staple' else 400)
        
        # 총 비용 (배송비 제외)
        cost_without_shipping = cv_p + cv_pr + cv_pl + cv_c + in_p + in_pr + in_pl + bind
        
        # 마진 적용 (배송비 제외)
        supply_cost = int(cost_without_shipping * (1 + margin / 100))
        
        # 부가세는 공금가 + 배송비에 적용
        vat = int((supply_cost + ship_cost) * 0.1)
        total_price = int((supply_cost + ship_cost + vat) / 10) * 10
        
        print(f"\n[BOOK 최종결과]")
        print(f"  - 표지 종이비: {cv_p}원, 인쇄비: {cv_pr}원, 판비: {cv_pl}원, 코팅비: {cv_c}원")
        print(f"  - 내지 종이비: {in_p}원, 인쇄비: {in_pr}원, 판비: {in_pl}원")
        print(f"  - 제본비: {bind}원")
        print(f"  - 배송비: {ship_cost}원")
        print(f"  - 소계(배송비제외): {cost_without_shipping}원")
        print(f"  - 공금가: {supply_cost}원 (마진 {margin}% 적용)")
        print(f"  - 부가세: {vat}원")
        print(f"  - 최종가: {total_price}원")
        
        return {
            'category': category,
            'qty': qty,
            'supply_cost': supply_cost,
            'vat': vat,
            'shipping': ship_cost,
            'total': total_price,
            'breakdown': {
                'cover': {'paper': cv_p, 'print': cv_pr, 'plate': cv_pl, 'coat': cv_c},
                'inner': {'paper': in_p, 'print': in_pr, 'plate': in_pl},
                'binding': bind,
                'shipping': ship_cost
            }
        }
    except Exception as e:
        print(f"❌ 책자 계산 오류: {e}")
        raise


# ========== 새로운 가격 관리 API (통합) ==========

@app.route('/api/admin/pricing', methods=['GET'])
@token_required
def get_pricing(current_user):
    """통합 가격 관리 데이터 조회"""
    try:
        result = {
            'paper_prices': [],
            'printing_costs': {},
            'plate_costs': {},
            'margin_settings': {},
            'binding_costs': {},
            'additional_costs': []
        }
        
        # 1. 종이 가격
        papers = PaperPrice.query.all()
        result['paper_prices'] = [p.to_dict() for p in papers]
        print(f"\n[GET /api/admin/pricing] 종이 가격: {len(papers)}개")
        for p in papers:
            print(f"  - {p.paper_type} {p.gram}g: 국전지={p.kook_price}, 46전지={p.sheet_4x6_price}")
        
        # 2. 인쇄비 (카테고리별)
        printing = PrintingCost.query.all()
        for p in printing:
            result['printing_costs'][p.category] = p.to_dict()
        print(f"[GET /api/admin/pricing] 인쇄비: {len(printing)}개")
        for p in printing:
            print(f"  - {p.category}: {p.to_dict()}")
        
        # 3. 판비 (대량옵셋만)
        plates = PlateCostNew.query.all()
        for p in plates:
            if p.category not in result['plate_costs']:
                result['plate_costs'][p.category] = []
            result['plate_costs'][p.category].append(p.to_dict())
        print(f"[GET /api/admin/pricing] 판비: {len(plates)}개")
        for p in plates:
            print(f"  - {p.category}: {p.to_dict()}")
        
        # 4. 마진율 (카테고리 & 회원별)
        margins = MarginSetting.query.all()
        for m in margins:
            key = m.category
            if key not in result['margin_settings']:
                result['margin_settings'][key] = {}
            result['margin_settings'][key][m.member_type] = m.margin_rate
        print(f"[GET /api/admin/pricing] 마진율: {len(margins)}개")
        for m in margins:
            print(f"  - {m.category} ({m.member_type}): {m.margin_rate}%")
        
        # 5. 제본비 (카테고리 & 바인딩별)
        bindings = BindingCost.query.all()
        for b in bindings:
            key = f"{b.category}_{b.binding_type}"
            if key not in result['binding_costs']:
                result['binding_costs'][key] = []
            result['binding_costs'][key].append(b.to_dict())
        print(f"[GET /api/admin/pricing] 제본비: {len(bindings)}개")
        for b in bindings:
            print(f"  - {b.category} ({b.binding_type}): {b.cost}원")
        
        # 6. 추가비용
        additional = AdditionalCost.query.all()
        result['additional_costs'] = [a.to_dict() for a in additional]
        print(f"[GET /api/admin/pricing] 추가비용: {len(additional)}개")
        for a in additional:
            print(f"  - {a.cost_name}: {a.cost}원")
        
        return jsonify({'success': True, 'data': result})
    
    except Exception as e:
        print(f"❌ 가격 데이터 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/pricing/paper/<int:paper_id>', methods=['PUT'])
@token_required
def update_paper_price(current_user, paper_id):
    """종이 가격 수정"""
    try:
        paper = PaperPrice.query.get(paper_id)
        if not paper:
            return jsonify({'success': False, 'message': '종이를 찾을 수 없습니다.'}), 404
        
        data = request.json
        
        if 'kook_price' in data:
            paper.kook_price = float(data['kook_price'])
        if 'sheet_4x6_price' in data:
            paper.sheet_4x6_price = float(data['sheet_4x6_price'])
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '종이 가격이 저장되었습니다.',
            'paper': paper.to_dict()
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 종이 가격 수정 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/pricing/printing/<category>', methods=['PUT'])
@token_required
def update_printing_cost(current_user, category):
    """인쇄비 수정"""
    try:
        data = request.json
        cover_cost = float(data.get('cover_print_cost', 0))
        inner_cost = float(data.get('inner_print_cost', 0))
        
        # PrintingCost 테이블에 저장
        printing = PrintingCost.query.filter_by(category=category).first()
        if not printing:
            printing = PrintingCost(
                category=category,
                cover_print_cost=cover_cost,
                inner_print_cost=inner_cost
            )
            db.session.add(printing)
        else:
            printing.cover_print_cost = cover_cost
            printing.inner_print_cost = inner_cost
        
        db.session.flush()
        
        # AdditionalCost 테이블에도 저장 (백엔드 계산에서 사용)
        # category에 따라 적절한 cost_name 매핑
        cost_name_mapping = {
            'digital': 'print_digital',
            'offset': 'print_offset',
            'indigo': 'print_indigo_color',  # cover는 color
            'indigo_mono': 'print_indigo_mono'
        }
        
        cost_name = cost_name_mapping.get(category)
        if cost_name:
            # cover 비용으로 저장
            additional = AdditionalCost.query.filter_by(cost_name=cost_name).first()
            if not additional:
                additional = AdditionalCost(
                    cost_name=cost_name,
                    unit='원',
                    cost=cover_cost
                )
                db.session.add(additional)
            else:
                additional.cost = cover_cost
            
            db.session.flush()
            
            # inner 비용을 별도로 저장 (mono 버전이 있으면)
            if category == 'indigo':
                mono_cost_name = 'print_indigo_mono'
                mono_additional = AdditionalCost.query.filter_by(cost_name=mono_cost_name).first()
                if not mono_additional:
                    mono_additional = AdditionalCost(
                        cost_name=mono_cost_name,
                        unit='원',
                        cost=inner_cost
                    )
                    db.session.add(mono_additional)
                else:
                    mono_additional.cost = inner_cost
        
        db.session.commit()
        
        print(f'[UPDATE_PRINTING_COST] {category}: cover={cover_cost}, inner={inner_cost}')
        print(f'[UPDATE_PRINTING_COST] AdditionalCost 테이블에도 저장됨')
        
        return jsonify({
            'success': True,
            'message': f'{category} 인쇄비가 저장되었습니다. (DB 및 AdditionalCost)',
            'printing': printing.to_dict()
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 인쇄비 수정 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/pricing/plate/<int:plate_id>', methods=['PUT'])
@token_required
def update_plate_cost(current_user, plate_id):
    """판비 수정"""
    try:
        plate = PlateCostNew.query.get(plate_id)
        if not plate:
            return jsonify({'success': False, 'message': '판비를 찾을 수 없습니다.'}), 404
        
        data = request.json
        
        if 'cost' in data:
            plate.cost = float(data['cost'])
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '판비가 저장되었습니다.',
            'plate': plate.to_dict()
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 판비 수정 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/pricing/margin', methods=['PUT'])
@token_required
@admin_required
def update_margin(current_user):
    """마진율 수정"""
    try:
        data = request.json
        category = data.get('category')
        member_type = data.get('member_type')  # 'general', 'business'
        margin_rate = float(data.get('margin_rate', 0))
        
        margin = MarginSetting.query.filter_by(
            category=category, 
            member_type=member_type
        ).first()
        
        if not margin:
            # 새로 생성
            margin = MarginSetting(
                category=category,
                member_type=member_type,
                margin_rate=margin_rate
            )
            db.session.add(margin)
        else:
            margin.margin_rate = margin_rate
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{category} ({member_type}) 마진율이 저장되었습니다.',
            'margin': margin.to_dict()
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 마진율 수정 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/pricing/binding', methods=['PUT'])
@token_required
def update_binding_cost(current_user):
    """제본비 수정"""
    try:
        data = request.json
        category = data.get('category')
        binding_type = data.get('binding_type')  # 'staple', 'perfect'
        cost = float(data.get('cost', 0))
        min_qty = data.get('min_qty', 0)  # 옵셋의 조건부 가격
        
        binding = BindingCost.query.filter_by(
            category=category,
            binding_type=binding_type,
            min_qty=min_qty
        ).first()
        
        if not binding:
            # 새로 생성
            binding = BindingCost(
                category=category,
                binding_type=binding_type,
                cost=cost,
                min_qty=min_qty
            )
            db.session.add(binding)
        else:
            binding.cost = cost
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{category} ({binding_type}) 제본비가 저장되었습니다.',
            'binding': binding.to_dict()
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 제본비 수정 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/pricing/additional/<int:cost_id>', methods=['PUT'])
@token_required
def update_additional_cost(current_user, cost_id):
    """추가비용 수정"""
    try:
        additional = AdditionalCost.query.get(cost_id)
        if not additional:
            return jsonify({'success': False, 'message': '추가비용을 찾을 수 없습니다.'}), 404
        
        data = request.json
        
        if 'cost' in data:
            additional.cost = float(data['cost'])
        if 'description' in data:
            additional.description = data['description']
        if 'unit' in data:
            additional.unit = data['unit']
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '추가비용이 저장되었습니다.',
            'additional': additional.to_dict()
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 추가비용 수정 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ========== NEW API: 카테고리/상품 관리 (이중 시스템) ==========

# ===== 카테고리 API =====
@app.route('/api/categories', methods=['GET'])
def get_categories():
    """모든 카테고리 조회 (계층 구조 지원)"""
    try:
        category_type = request.args.get('type')  # 'quote', 'sellable', 또는 전체
        parent_only = request.args.get('parent_only', 'false').lower() == 'true'
        
        query = Category.query
        
        if category_type:
            query = query.filter_by(category_type=category_type)
        
        if parent_only:
            # 부모 카테고리만 반환 (children 포함)
            query = query.filter_by(parent_id=None)
        
        categories = query.order_by(Category.display_order).all()
        
        return jsonify({
            'success': True,
            'data': [cat.to_dict() for cat in categories]
        })
    except Exception as e:
        print(f"❌ 카테고리 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/categories', methods=['POST'])
@token_required
def create_category(current_user):
    """카테고리 추가 (계층 구조 지원)"""
    try:
        # 관리자 권한 확인
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        data = request.json
        name = data.get('name')
        category_type = data.get('category_type')  # 'quote' or 'sellable'
        parent_id = data.get('parent_id')  # 선택사항: 부모 카테고리 ID
        
        if not name or not category_type:
            return jsonify({'success': False, 'message': '필수 항목이 누락되었습니다.'}), 400
        
        # parent_id 검증
        if parent_id:
            parent = Category.query.get(parent_id)
            if not parent:
                return jsonify({'success': False, 'message': '부모 카테고리가 존재하지 않습니다.'}), 404
        
        # 중복 확인 (동일 부모 내에서만)
        existing = Category.query.filter_by(name=name, parent_id=parent_id).first()
        if existing:
            return jsonify({'success': False, 'message': '이미 존재하는 카테고리입니다.'}), 409
        
        # 다음 표시 순서 계산 (부모별로)
        if parent_id:
            max_order = db.session.query(db.func.max(Category.display_order)).filter_by(parent_id=parent_id).scalar() or 0
        else:
            max_order = db.session.query(db.func.max(Category.display_order)).filter_by(parent_id=None).scalar() or 0
        
        category = Category(
            name=name,
            description=data.get('description', ''),
            category_type=category_type,
            icon=data.get('icon', ''),
            parent_id=parent_id,
            display_order=max_order + 1,
            is_active=data.get('is_active', True)
        )
        
        db.session.add(category)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{name} 카테고리가 생성되었습니다.',
            'data': category.to_dict()
        }), 201
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 카테고리 생성 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/categories/<int:cat_id>', methods=['PUT'])
@token_required
def update_category(cat_id, current_user):
    """카테고리 수정 (계층 구조 지원)"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        category = Category.query.get(cat_id)
        if not category:
            return jsonify({'success': False, 'message': '카테고리를 찾을 수 없습니다.'}), 404
        
        data = request.json
        
        if 'name' in data:
            category.name = data['name']
        if 'description' in data:
            category.description = data['description']
        if 'icon' in data:
            category.icon = data['icon']
        if 'is_active' in data:
            category.is_active = data['is_active']
        if 'parent_id' in data:
            parent_id = data['parent_id']
            if parent_id:
                parent = Category.query.get(parent_id)
                if not parent:
                    return jsonify({'success': False, 'message': '부모 카테고리가 존재하지 않습니다.'}), 404
                if parent_id == cat_id:
                    return jsonify({'success': False, 'message': '자신을 부모로 설정할 수 없습니다.'}), 400
            category.parent_id = parent_id
        
        category.updated_at = now_kst()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '카테고리가 업데이트되었습니다.',
            'data': category.to_dict()
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 카테고리 수정 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/categories/<int:cat_id>', methods=['DELETE'])
@token_required
def delete_category(cat_id, current_user):
    """카테고리 삭제 (계층 구조 안전 삭제)"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        category = Category.query.get(cat_id)
        if not category:
            return jsonify({'success': False, 'message': '카테고리를 찾을 수 없습니다.'}), 404
        
        # 자식 카테고리 확인
        if category.children:
            return jsonify({
                'success': False,
                'message': f'이 카테고리에는 {len(category.children)}개의 하위 카테고리가 있습니다. 먼저 하위 카테고리를 삭제해주세요.'
            }), 409
        
        # 하위 상품 확인
        if category.products:
            return jsonify({
                'success': False,
                'message': f'이 카테고리에는 {len(category.products)}개의 상품이 있습니다. 먼저 상품을 삭제해주세요.'
            }), 409
        
        db.session.delete(category)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '카테고리가 삭제되었습니다.'
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 카테고리 삭제 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/categories/reorder', methods=['PATCH'])
@token_required
def reorder_categories(current_user):
    """카테고리 순서 변경"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        data = request.json
        category_ids = data.get('category_ids', [])  # 순서대로 정렬된 ID 배열
        
        for order, cat_id in enumerate(category_ids):
            category = Category.query.get(cat_id)
            if category:
                category.display_order = order
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '카테고리 순서가 변경되었습니다.'
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 카테고리 순서 변경 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ===== 상품 API =====
@app.route('/api/products', methods=['GET'])
def get_products():
    """상품 목록 조회"""
    try:
        category_id = request.args.get('category_id', type=int)
        product_type = request.args.get('type')  # 'quote_based', 'sellable'
        
        query = Product.query
        
        if category_id:
            query = query.filter_by(category_id=category_id)
        
        if product_type:
            query = query.filter_by(product_type=product_type)
        
        products = query.order_by(Product.display_order).all()
        
        return jsonify({
            'success': True,
            'data': [prod.to_dict() for prod in products]
        })
    
    except Exception as e:
        print(f"❌ 상품 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/products', methods=['POST'])
@token_required
def create_product(current_user):
    """상품 추가"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        data = request.json
        category_id = data.get('category_id')
        name = data.get('name')
        product_type = data.get('product_type')
        
        if not all([category_id, name, product_type]):
            return jsonify({'success': False, 'message': '필수 항목이 누락되었습니다.'}), 400
        
        # 카테고리 확인
        category = Category.query.get(category_id)
        if not category:
            return jsonify({'success': False, 'message': '카테고리를 찾을 수 없습니다.'}), 404
        
        # 다음 표시 순서 계산
        max_order = db.session.query(db.func.max(Product.display_order)).filter_by(category_id=category_id).scalar() or 0
        
        product = Product(
            category_id=category_id,
            name=name,
            description=data.get('description', ''),
            product_type=product_type,
            margin=data.get('margin', 0),
            image_url=data.get('image_url', ''),
            is_active=data.get('is_active', True),
            display_order=max_order + 1
        )
        
        if product_type == 'quote_based':
            product.quote_settings = json.dumps(data.get('quote_settings', {}))
        else:  # sellable
            product.fixed_price = data.get('fixed_price', 0)
            product.quantity = data.get('quantity', 0)
            product.stock_alert = data.get('stock_alert', 10)
            product.cost_price = data.get('cost_price', 0)
            product.sellable_specs = json.dumps(data.get('sellable_specs', {}))
        
        db.session.add(product)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{name} 상품이 생성되었습니다.',
            'data': product.to_dict()
        }), 201
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 상품 생성 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/products/<int:product_id>', methods=['PUT'])
@token_required
def update_product(product_id, current_user):
    """상품 수정"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        product = Product.query.get(product_id)
        if not product:
            return jsonify({'success': False, 'message': '상품을 찾을 수 없습니다.'}), 404
        
        data = request.json
        
        if 'name' in data:
            product.name = data['name']
        if 'description' in data:
            product.description = data['description']
        if 'margin' in data:
            product.margin = data['margin']
        if 'image_url' in data:
            product.image_url = data['image_url']
        if 'is_active' in data:
            product.is_active = data['is_active']
        
        # 타입별 추가 필드
        if product.product_type == 'quote_based':
            if 'quote_settings' in data:
                product.quote_settings = json.dumps(data['quote_settings'])
        else:  # sellable
            if 'fixed_price' in data:
                product.fixed_price = data['fixed_price']
            if 'quantity' in data:
                product.quantity = data['quantity']
            if 'stock_alert' in data:
                product.stock_alert = data['stock_alert']
            if 'cost_price' in data:
                product.cost_price = data['cost_price']
            if 'sellable_specs' in data:
                product.sellable_specs = json.dumps(data['sellable_specs'])
        
        product.updated_at = now_kst()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '상품이 업데이트되었습니다.',
            'data': product.to_dict()
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 상품 수정 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/products/<int:product_id>', methods=['DELETE'])
@token_required
def delete_product(product_id, current_user):
    """상품 삭제"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        product = Product.query.get(product_id)
        if not product:
            return jsonify({'success': False, 'message': '상품을 찾을 수 없습니다.'}), 404
        
        db.session.delete(product)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '상품이 삭제되었습니다.'
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 상품 삭제 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/products/reorder', methods=['PATCH'])
@token_required
def reorder_products(current_user):
    """상품 순서 변경"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        data = request.json
        product_ids = data.get('product_ids', [])
        
        for order, prod_id in enumerate(product_ids):
            product = Product.query.get(prod_id)
            if product:
                product.display_order = order
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '상품 순서가 변경되었습니다.'
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 상품 순서 변경 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ===== 상품 바리언트 API =====
@app.route('/api/products/<int:product_id>/variants', methods=['GET'])
def get_product_variants(product_id):
    """상품 바리언트 조회"""
    try:
        variants = ProductVariant.query.filter_by(product_id=product_id).order_by(ProductVariant.display_order).all()
        
        return jsonify({
            'success': True,
            'data': [var.to_dict() for var in variants]
        })
    
    except Exception as e:
        print(f"❌ 바리언트 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/products/<int:product_id>/variants', methods=['POST'])
@token_required
def create_product_variant(product_id, current_user):
    """상품 바리언트 추가"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        product = Product.query.get(product_id)
        if not product:
            return jsonify({'success': False, 'message': '상품을 찾을 수 없습니다.'}), 404
        
        data = request.json
        
        variant = ProductVariant(
            product_id=product_id,
            binding_type=data.get('binding_type'),
            guide_text=data.get('guide_text', ''),
            ship_info=data.get('ship_info', ''),
            info_html=data.get('info_html', ''),
            variant_price=data.get('variant_price', 0),
            variant_specs=json.dumps(data.get('variant_specs', {}))
        )
        
        db.session.add(variant)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '바리언트가 생성되었습니다.',
            'data': variant.to_dict()
        }), 201
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 바리언트 생성 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/products/<int:product_id>/variants/<int:variant_id>', methods=['PUT'])
@token_required
def update_product_variant(product_id, variant_id, current_user):
    """상품 바리언트 수정"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        variant = ProductVariant.query.filter_by(id=variant_id, product_id=product_id).first()
        if not variant:
            return jsonify({'success': False, 'message': '바리언트를 찾을 수 없습니다.'}), 404
        
        data = request.json
        
        if 'guide_text' in data:
            variant.guide_text = data['guide_text']
        if 'ship_info' in data:
            variant.ship_info = data['ship_info']
        if 'info_html' in data:
            variant.info_html = data['info_html']
        if 'variant_price' in data:
            variant.variant_price = data['variant_price']
        if 'variant_specs' in data:
            variant.variant_specs = json.dumps(data['variant_specs'])
        
        variant.updated_at = now_kst()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '바리언트가 업데이트되었습니다.',
            'data': variant.to_dict()
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 바리언트 수정 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/products/<int:product_id>/variants/<int:variant_id>', methods=['DELETE'])
@token_required
def delete_product_variant(product_id, variant_id, current_user):
    """상품 바리언트 삭제"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        variant = ProductVariant.query.filter_by(id=variant_id, product_id=product_id).first()
        if not variant:
            return jsonify({'success': False, 'message': '바리언트를 찾을 수 없습니다.'}), 404
        
        db.session.delete(variant)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '바리언트가 삭제되었습니다.'
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 바리언트 삭제 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ===== 판매형 상품 옵션 API =====
@app.route('/api/products/<int:product_id>/options', methods=['GET'])
def get_product_options(product_id):
    """상품 옵션 조회"""
    try:
        options = SellableProductOption.query.filter_by(product_id=product_id).order_by(SellableProductOption.display_order).all()
        
        return jsonify({
            'success': True,
            'data': [opt.to_dict() for opt in options]
        })
    
    except Exception as e:
        print(f"❌ 옵션 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/products/<int:product_id>/options', methods=['POST'])
@token_required
def create_product_option(product_id, current_user):
    """상품 옵션 추가"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        product = Product.query.get(product_id)
        if not product:
            return jsonify({'success': False, 'message': '상품을 찾을 수 없습니다.'}), 404
        
        if product.product_type != 'sellable':
            return jsonify({'success': False, 'message': '판매형 상품에만 옵션을 추가할 수 있습니다.'}), 400
        
        data = request.json
        option_name = data.get('option_name')
        option_values = data.get('option_values', [])
        
        if not option_name or not option_values:
            return jsonify({'success': False, 'message': '필수 항목이 누락되었습니다.'}), 400
        
        # 다음 표시 순서
        max_order = db.session.query(db.func.max(SellableProductOption.display_order)).filter_by(product_id=product_id).scalar() or 0
        
        option = SellableProductOption(
            product_id=product_id,
            option_name=option_name,
            option_values=json.dumps(option_values),
            is_required=data.get('is_required', False),
            display_order=max_order + 1
        )
        
        db.session.add(option)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{option_name} 옵션이 생성되었습니다.',
            'data': option.to_dict()
        }), 201
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 옵션 생성 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/products/<int:product_id>/options/<int:option_id>', methods=['PUT'])
@token_required
def update_product_option(product_id, option_id, current_user):
    """상품 옵션 수정"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        option = SellableProductOption.query.filter_by(id=option_id, product_id=product_id).first()
        if not option:
            return jsonify({'success': False, 'message': '옵션을 찾을 수 없습니다.'}), 404
        
        data = request.json
        
        if 'option_name' in data:
            option.option_name = data['option_name']
        if 'option_values' in data:
            option.option_values = json.dumps(data['option_values'])
        if 'is_required' in data:
            option.is_required = data['is_required']
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '옵션이 업데이트되었습니다.',
            'data': option.to_dict()
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 옵션 수정 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/products/<int:product_id>/options/<int:option_id>', methods=['DELETE'])
@token_required
def delete_product_option(product_id, option_id, current_user):
    """상품 옵션 삭제"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        option = SellableProductOption.query.filter_by(id=option_id, product_id=product_id).first()
        if not option:
            return jsonify({'success': False, 'message': '옵션을 찾을 수 없습니다.'}), 404
        
        db.session.delete(option)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '옵션이 삭제되었습니다.'
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 옵션 삭제 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ===== 상품 엑셀 대량 관리 API =====

@app.route('/api/products/export/excel', methods=['GET'])
@token_required
def export_products_excel(current_user):
    """상품 데이터를 엑셀로 내보내기"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        # 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "상품"
        
        # 헤더 설정
        headers = ['상품명', '기본가격', '카테고리', '마진율(%)', '옵션명1', '옵션값1', '추가요금1', 
                   '옵션명2', '옵션값2', '추가요금2', '재고', '관리코드', '사용여부']
        ws.append(headers)
        
        # 헤더 스타일
        header_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
        header_font = Font(bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 데이터 행 추가
        products = Product.query.all()
        
        for product in products:
            category = Category.query.get(product.category_id)
            category_name = category.name if category else ''
            
            # 바리언트 조회
            variants = ProductVariant.query.filter_by(product_id=product.id).all()
            
            if not variants:
                # 바리언트가 없으면 기본 행만 추가
                row = [
                    product.name,
                    product.fixed_price or 0,
                    category_name,
                    product.margin or 0,
                    '', '', '',  # 옵션1
                    '', '', '',  # 옵션2
                    product.quantity or 0,
                    product.sku or '',
                    'Y' if product.is_active else 'N'
                ]
                ws.append(row)
            else:
                # 각 바리언트별로 행 추가
                for variant in variants:
                    # 바리언트의 옵션 파싱
                    variant_specs = {}
                    try:
                        if variant.variant_specs:
                            variant_specs = json.loads(variant.variant_specs)
                    except:
                        pass
                    
                    # 옵션 정보 추출 (최대 2개)
                    option_info = []
                    for key, value in variant_specs.items():
                        option_info.append((key, value, 0))  # 추가요금은 따로 계산
                        if len(option_info) >= 2:
                            break
                    
                    # 옵션 채우기
                    while len(option_info) < 2:
                        option_info.append(('', '', 0))
                    
                    row = [
                        product.name,
                        product.fixed_price or 0,
                        category_name,
                        product.margin or 0,
                        option_info[0][0],
                        option_info[0][1],
                        option_info[0][2],
                        option_info[1][0],
                        option_info[1][1],
                        option_info[1][2],
                        variant.quantity or 0,
                        variant.sku or '',
                        'Y' if variant.is_active else 'N'
                    ]
                    ws.append(row)
        
        # 열 너비 설정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        for col in ['E', 'F', 'G', 'H', 'I', 'J']:
            ws.column_dimensions[col].width = 12
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 8
        
        # 엑셀 파일 저장
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='products_export.xlsx'
        )
    
    except Exception as e:
        print(f"❌ 엑셀 내보내기 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/products/import/excel', methods=['POST'])
@token_required
def import_products_excel(current_user):
    """엑셀에서 상품 데이터 대량 가져오기"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        # 파일 확인
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '파일이 없습니다.'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '파일을 선택해주세요.'}), 400
        
        # 파일 읽기
        from openpyxl import load_workbook
        import csv
        from io import StringIO
        
        errors = []
        success_count = 0
        
        try:
            # Excel 파일 읽기
            if file.filename.endswith('.xlsx') or file.filename.endswith('.xls'):
                wb = load_workbook(file)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                
                if len(rows) < 2:
                    return jsonify({'success': False, 'message': '헤더 행만 있습니다.'}), 400
                
                # 헤더 파싱
                headers = rows[0]
                
                # 데이터 행 처리
                for row_idx, row in enumerate(rows[1:], start=2):
                    try:
                        # 필수 필드 확인
                        product_name = row[0] if row[0] else None
                        base_price = float(row[1]) if row[1] else 0
                        category_name = row[2] if row[2] else None
                        margin = float(row[3]) if row[3] else 0
                        
                        if not product_name or not category_name:
                            errors.append(f"행 {row_idx}: 상품명과 카테고리는 필수입니다.")
                            continue
                        
                        # 카테고리 조회
                        category = Category.query.filter_by(name=category_name).first()
                        if not category:
                            errors.append(f"행 {row_idx}: '{category_name}' 카테고리가 없습니다.")
                            continue
                        
                        # 상품 조회 또는 생성
                        product = Product.query.filter_by(name=product_name, category_id=category.id).first()
                        
                        if not product:
                            product = Product(
                                name=product_name,
                                category_id=category.id,
                                product_type='sellable',
                                fixed_price=base_price,
                                margin=margin,
                                is_active=True
                            )
                            db.session.add(product)
                            db.session.flush()
                        else:
                            product.fixed_price = base_price
                            product.margin = margin
                        
                        # 옵션 정보 추출
                        variant_specs = {}
                        option_prices = {}
                        
                        # 옵션1
                        if row[4] and row[5]:  # 옵션명과 옵션값
                            variant_specs[row[4]] = row[5]
                            option_prices[row[4]] = float(row[6]) if row[6] else 0
                        
                        # 옵션2
                        if row[7] and row[8]:  # 옵션명과 옵션값
                            variant_specs[row[7]] = row[8]
                            option_prices[row[7]] = float(row[9]) if row[9] else 0
                        
                        stock = int(row[10]) if row[10] else 0
                        sku = row[11] if row[11] else ''
                        is_active = str(row[12]).upper() == 'Y' if row[12] else True
                        
                        # 바리언트 생성 또는 업데이트
                        variant_key = json.dumps(variant_specs, sort_keys=True, ensure_ascii=False)
                        variant = ProductVariant.query.filter_by(
                            product_id=product.id,
                            variant_specs=variant_key
                        ).first()
                        
                        if not variant:
                            variant = ProductVariant(
                                product_id=product.id,
                                variant_specs=variant_key,
                                variant_price=base_price + sum(option_prices.values()),
                                quantity=stock,
                                sku=sku,
                                is_active=is_active
                            )
                            db.session.add(variant)
                        else:
                            variant.variant_price = base_price + sum(option_prices.values())
                            variant.quantity = stock
                            variant.sku = sku
                            variant.is_active = is_active
                        
                        success_count += 1
                    
                    except Exception as row_err:
                        errors.append(f"행 {row_idx}: {str(row_err)}")
            
            # CSV 파일 읽기
            elif file.filename.endswith('.csv'):
                content = file.read().decode('utf-8')
                reader = csv.DictReader(StringIO(content))
                
                for row_idx, row in enumerate(reader, start=2):
                    try:
                        product_name = row.get('상품명')
                        base_price = float(row.get('기본가격', 0))
                        category_name = row.get('카테고리')
                        margin = float(row.get('마진율(%)', 0))
                        
                        if not product_name or not category_name:
                            errors.append(f"행 {row_idx}: 상품명과 카테고리는 필수입니다.")
                            continue
                        
                        # 나머지 처리는 동일...
                        success_count += 1
                    
                    except Exception as row_err:
                        errors.append(f"행 {row_idx}: {str(row_err)}")
            
            else:
                return jsonify({'success': False, 'message': '.xlsx, .xls, .csv 파일만 지원합니다.'}), 400
            
            # 데이터 저장
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f'{success_count}개 상품이 저장되었습니다.',
                'success_count': success_count,
                'errors': errors
            })
        
        except Exception as parse_err:
            db.session.rollback()
            return jsonify({'success': False, 'message': f'파일 파싱 오류: {str(parse_err)}'}), 400
    
    except Exception as e:
        db.session.rollback()
        print(f"❌ 엑셀 가져오기 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/products/template/excel', methods=['GET'])
@token_required
def download_product_template(current_user):
    """상품 입력 템플릿 다운로드"""
    try:
        user = User.query.filter_by(user_id=current_user).first()
        if not user or user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        # 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "상품"
        
        # 헤더 설정
        headers = ['상품명', '기본가격', '카테고리', '마진율(%)', '옵션명1', '옵션값1', '추가요금1', 
                   '옵션명2', '옵션값2', '추가요금2', '재고', '관리코드', '사용여부']
        ws.append(headers)
        
        # 헤더 스타일
        header_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
        header_font = Font(bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 샘플 데이터
        sample_rows = [
            ['티셔츠', 10000, '의류', 10, '크기', 'S', 0, '색상', '빨강', 0, 100, 'SKU001', 'Y'],
            ['티셔츠', 10000, '의류', 10, '크기', 'S', 0, '색상', '파랑', 150, 50, 'SKU002', 'Y'],
            ['티셔츠', 10000, '의류', 10, '크기', 'M', 200, '색상', '빨강', 0, 80, 'SKU003', 'Y'],
            ['티셔츠', 10000, '의류', 10, '크기', 'M', 200, '색상', '파랑', 150, 60, 'SKU004', 'Y'],
        ]
        
        for row in sample_rows:
            ws.append(row)
        
        # 샘플 행 스타일 (연한 파란색)
        sample_fill = PatternFill(start_color='D9E8F5', end_color='D9E8F5', fill_type='solid')
        for row_idx in range(2, len(sample_rows) + 2):
            for cell in ws[row_idx]:
                cell.fill = sample_fill
        
        # 설명 행 추가
        ws.append([])
        ws.append(['※ 사용 방법:'])
        ws.append(['1. 상품명, 기본가격, 카테고리, 마린율은 필수입니다.'])
        ws.append(['2. 옵션명과 옵션값을 입력하면 자동으로 조합 가격이 계산됩니다.'])
        ws.append(['3. 추가요금은 기본가격에 더해지는 금액입니다. (예: 크기M은 +200원)'])
        ws.append(['4. 같은 상품의 다른 옵션 조합은 새로운 행으로 입력합니다.'])
        ws.append(['5. 카테고리는 미리 생성된 카테고리명을 정확히 입력해야 합니다.'])
        ws.append(['6. 사용여부는 Y 또는 N으로 입력합니다.'])
        
        # 설명 텍스트 스타일
        for row_idx in range(len(sample_rows) + 3, len(sample_rows) + 9):
            for cell in ws[row_idx]:
                cell.font = Font(size=9, italic=True)
                cell.alignment = Alignment(wrap_text=True)
        
        # 열 너비 설정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        for col in ['E', 'F', 'G', 'H', 'I', 'J']:
            ws.column_dimensions[col].width = 12
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 8
        
        # 엑셀 파일 저장
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='product_template.xlsx'
        )
    
    except Exception as e:
        print(f"❌ 템플릿 다운로드 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ========== 주문 내역서 JPG 다운로드 ==========
@app.route('/api/admin/orders/invoice/download', methods=['POST'])
def download_order_invoices():
    """선택된 주문들의 내역서를 JPG로 다운로드"""
    try:
        # 로컬 환경 감지
        is_local = request.host.split(':')[0] in ['localhost', '127.0.0.1']
        
        # 토큰 인증 확인
        current_user = None
        if 'Authorization' in request.headers:
            try:
                auth_header = request.headers.get('Authorization', '')
                token = auth_header.split(' ')[1] if ' ' in auth_header else None
                if token:
                    user_id = verify_token(token)
                    if user_id:
                        current_user = User.query.filter_by(user_id=user_id).first()
            except:
                pass
        
        # 로컬 환경에서 토큰 없을 때 테스트 관리자 자동 생성
        if not current_user:
            if is_local:
                print("[download_order_invoices] 🔧 로컬 환경 - 테스트 관리자 사용")
                test_admin = User.query.filter_by(user_id='admin').first()
                if not test_admin:
                    test_admin = User(
                        user_id='admin',
                        password='admin123',
                        role='admin',
                        name='테스트 관리자'
                    )
                    db.session.add(test_admin)
                    db.session.commit()
                current_user = test_admin
            else:
                return jsonify({'success': False, 'message': '인증이 필요합니다.'}), 401
        
        # 관리자 역할 확인
        if current_user.role != 'admin':
            return jsonify({'success': False, 'message': '관리자 권한이 필요합니다.'}), 403
        
        data = request.json
        order_ids = data.get('order_ids', [])
        
        if not order_ids:
            return jsonify({'success': False, 'message': '주문을 선택해주세요.'}), 400
        
        print(f"[주문 내역서 다운로드] {len(order_ids)}개 주문 JPG 생성 시작")
        
        # 임시 디렉토리 생성
        with tempfile.TemporaryDirectory() as temp_dir:
            jpg_files = []
            
            for order_id in order_ids:
                order = Order.query.filter_by(order_id=order_id).first()
                if not order:
                    print(f"[WARNING] 주문을 찾을 수 없음: {order_id}")
                    continue
                
                try:
                    # HTML 생성
                    html_content = create_order_invoice_html(order)
                    
                    # JPG 파일 경로
                    temp_jpg = os.path.join(temp_dir, f'{order_id}.jpg')
                    
                    # HTML을 JPG로 변환
                    html_to_jpg(html_content, temp_jpg)
                    jpg_files.append((order_id, temp_jpg))
                    print(f"✅ JPG 생성 완료: {order_id}")
                    
                except Exception as e:
                    print(f"❌ {order_id} JPG 생성 실패: {e}")
                    continue
            
            if not jpg_files:
                return jsonify({'success': False, 'message': 'JPG를 생성할 수 없습니다.'}), 500
            
            # 1개 파일: 직접 다운로드
            if len(jpg_files) == 1:
                order_id, jpg_path = jpg_files[0]
                with open(jpg_path, 'rb') as f:
                    jpg_data = f.read()
                
                return send_file(
                    BytesIO(jpg_data),
                    mimetype='image/jpeg',
                    as_attachment=True,
                    download_name=f'주문내역서_{order_id}.jpg'
                )
            
            # 여러 파일: ZIP으로 압축
            else:
                zip_path = os.path.join(temp_dir, 'invoices.zip')
                with zipfile.ZipFile(zip_path, 'w') as zf:
                    for order_id, jpg_path in jpg_files:
                        zf.write(jpg_path, arcname=f'주문내역서_{order_id}.jpg')
                
                with open(zip_path, 'rb') as f:
                    zip_data = f.read()
                
                return send_file(
                    BytesIO(zip_data),
                    mimetype='application/zip',
                    as_attachment=True,
                    download_name=f'주문내역서_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
                )
    
    except Exception as e:
        print(f"[ERROR] 주문 내역서 다운로드 실패: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

if __name__ == '__main__':
    print("[*] Flask 서버를 시작합니다...")
    print("[*] URL: http://localhost:5000")
    
    # 앱 시작 시 상태값 마이그레이션 (한 번만 실행됨)
    with app.app_context():
        migrate_status_to_korean()
        
        # 라우트 확인
        print("\n[DEBUG] Registered routes:")
        for rule in app.url_map.iter_rules():
            if 'category' in rule.rule or 'calculate' in rule.rule or 'popup' in rule.rule:
                print(f"  {rule.rule} -> {rule.endpoint} {rule.methods}")
    
    app.run(host='0.0.0.0', port=5000, debug=True)
